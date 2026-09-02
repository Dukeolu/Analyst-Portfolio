"""
Case 03 -- build the Excel workbook deliverable.

This is the case's Excel showcase: live formulas throughout (no
hardcoded results), a conditional-formatting heatmap for the region x
category variance matrix, and a margin bridge that decomposes total
profit variance into volume, discount, and COGS effects -- three
components that are provably additive back to the total (see the
"Check" column in Monthly Detail).

Sheets:
  README                    context + color legend
  Monthly Detail            576 rows, budget vs actual, formula-derived variances
  Region x Category Summary flat table + pivoted heatmap matrix (SUMIFS)
  Margin Bridge             budget -> volume -> discount -> COGS -> actual, with chart
  Top Variance Drivers      worst 8 combos, ranked with SMALL + INDEX/MATCH
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[1]
RAW = BASE / "data" / "raw"
OUT = BASE / "excel" / "case03_margin_analysis.xlsx"

df = pd.read_csv(RAW / "budget_vs_actual.csv")
N = len(df)

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F2933")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT, color="0000FF", size=10)
FORMULA_FONT = Font(name=FONT, color="000000", size=10)
TITLE_FONT = Font(name=FONT, bold=True, size=15)
SUB_FONT = Font(name=FONT, size=10.5, italic=True, color="52514E")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = '$#,##0;($#,##0);"-"'
PCT1 = '0.0%'

wb = Workbook()

# =========================================================== README
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100
ws["A1"] = "Case 03 — Finding Where Margin Is Quietly Leaking"
ws["A1"].font = TITLE_FONT
ws["A2"] = "Finance & Budgeting case study — budget vs. actual profitability by region and category"
ws["A2"].font = SUB_FONT
lines = [
    "",
    "Business problem: the business is profitable overall, but budget-to-actual reviews keep",
    "coming in over on cost-to-serve in specific regions and categories. This workbook finds where.",
    "",
    "Sheet guide:",
    "  Monthly Detail             576 rows (4 regions x 6 categories x 24 months), budget vs. actual,",
    "                             with live formulas for revenue/discount/COGS variance and a 3-part",
    "                             profit-variance bridge (Volume, Discount, COGS effects).",
    "  Region x Category Summary  Flat SUMIFS table + a pivoted heatmap matrix flagging the leaks.",
    "  Margin Bridge              Total budget profit -> Volume -> Discount -> COGS -> actual profit,",
    "                             with a chart. All three effects are provably additive to the total",
    "                             variance (see the Check column in Monthly Detail).",
    "  Top Variance Drivers       Worst 8 region x category combos, ranked with SMALL + INDEX/MATCH.",
    "",
    "Color legend:",
    "  Blue text   = input data (as pulled from budget/actuals systems)",
    "  Black text  = formulas (recalculate if the inputs change)",
    "  Heatmap     = red = profit below budget, green = profit above budget",
    "",
    "Data note: this dataset is simulated (see the case README for why and how) -- built to be",
    "realistic, with two deliberate variance drivers seeded into an otherwise on-plan business:",
    "a ramping supplier cost overrun in Electronics Accessories, and discount creep in Outdoor &",
    "Sporting. Everything else moves with ordinary month-to-month noise around budget.",
]
for i, line in enumerate(lines, start=3):
    ws[f"A{i}"] = line
    ws[f"A{i}"].font = Font(name=FONT, size=10.5, bold=line.strip().endswith(":") and not line.startswith(" "))

# =========================================================== Monthly Detail
ws2 = wb.create_sheet("Monthly Detail")
headers = [
    "Month", "Region", "Category",
    "Budget Gross Rev", "Budget Discount %", "Budget COGS %",
    "Budget Net Rev", "Budget COGS $", "Budget Profit",
    "Actual Gross Rev", "Actual Discount %", "Actual COGS %",
    "Actual Net Rev", "Actual COGS $", "Actual Profit",
    "Profit Variance $", "Volume Effect $", "Discount Effect $", "COGS Effect $", "Check (should = 0)",
]
for c, h in enumerate(headers, start=1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws2.freeze_panes = "A2"

for i, row in enumerate(df.itertuples(index=False), start=2):
    ws2.cell(row=i, column=1, value=row.month).font = INPUT_FONT
    ws2.cell(row=i, column=2, value=row.region).font = INPUT_FONT
    ws2.cell(row=i, column=3, value=row.category).font = INPUT_FONT
    ws2.cell(row=i, column=4, value=row.budget_revenue).font = INPUT_FONT
    ws2.cell(row=i, column=5, value=row.budget_discount_pct).font = INPUT_FONT
    ws2.cell(row=i, column=6, value=row.budget_cogs_pct).font = INPUT_FONT
    # formulas
    ws2.cell(row=i, column=7, value=f"=D{i}*(1-E{i})").font = FORMULA_FONT
    ws2.cell(row=i, column=8, value=f"=G{i}*F{i}").font = FORMULA_FONT
    ws2.cell(row=i, column=9, value=f"=G{i}-H{i}").font = FORMULA_FONT

    ws2.cell(row=i, column=10, value=row.actual_revenue).font = INPUT_FONT
    ws2.cell(row=i, column=11, value=row.actual_discount_pct).font = INPUT_FONT
    ws2.cell(row=i, column=12, value=row.actual_cogs_pct).font = INPUT_FONT
    ws2.cell(row=i, column=13, value=f"=J{i}*(1-K{i})").font = FORMULA_FONT
    ws2.cell(row=i, column=14, value=f"=M{i}*L{i}").font = FORMULA_FONT
    ws2.cell(row=i, column=15, value=f"=M{i}-N{i}").font = FORMULA_FONT

    ws2.cell(row=i, column=16, value=f"=O{i}-I{i}").font = FORMULA_FONT
    # volume effect = (actual_gross - budget_gross) * (1-budget_disc) * (1-budget_cogs)
    ws2.cell(row=i, column=17, value=f"=(J{i}-D{i})*(1-E{i})*(1-F{i})").font = FORMULA_FONT
    # discount effect = actual_gross * (budget_disc - actual_disc) * (1-budget_cogs)
    ws2.cell(row=i, column=18, value=f"=J{i}*(E{i}-K{i})*(1-F{i})").font = FORMULA_FONT
    # cogs effect = actual_net_rev * (budget_cogs - actual_cogs)
    ws2.cell(row=i, column=19, value=f"=M{i}*(F{i}-L{i})").font = FORMULA_FONT
    # check: the three effects should sum exactly to the profit variance
    ws2.cell(row=i, column=20, value=f"=ROUND(P{i}-(Q{i}+R{i}+S{i}),2)").font = FORMULA_FONT

    for col in (5, 6, 11, 12):
        ws2.cell(row=i, column=col).number_format = PCT1
    for col in (4, 7, 8, 9, 10, 13, 14, 15, 16, 17, 18, 19, 20):
        ws2.cell(row=i, column=col).number_format = MONEY

widths = [9, 11, 22, 14, 12, 11, 13, 12, 13, 14, 12, 11, 13, 12, 13, 14, 13, 14, 12, 14]
for c, w in enumerate(widths, start=1):
    ws2.column_dimensions[get_column_letter(c)].width = w

last_row = N + 1

# =========================================================== Region x Category Summary
ws3 = wb.create_sheet("Region x Category Summary")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "Total Profit Variance by Region x Category (24 months)"
ws3["A1"].font = Font(name=FONT, bold=True, size=13)
ws3["A2"] = "Flat table (left) drives the heatmap matrix (below) via SUMIFS against Monthly Detail."
ws3["A2"].font = SUB_FONT

categories = sorted(df["category"].unique())
regions = sorted(df["region"].unique())

flat_headers = ["Category", "Region", "Budget Profit", "Actual Profit", "Variance $", "Variance %"]
for c, h in enumerate(flat_headers, start=1):
    cell = ws3.cell(row=4, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL

flat_row = 5
combo_rows = {}
for cat in categories:
    for reg in regions:
        ws3.cell(row=flat_row, column=1, value=cat).font = FORMULA_FONT
        ws3.cell(row=flat_row, column=2, value=reg).font = FORMULA_FONT
        ws3.cell(row=flat_row, column=3,
                 value=f"=SUMIFS('Monthly Detail'!$I$2:$I${last_row},'Monthly Detail'!$C$2:$C${last_row},A{flat_row},'Monthly Detail'!$B$2:$B${last_row},B{flat_row})").font = FORMULA_FONT
        ws3.cell(row=flat_row, column=4,
                 value=f"=SUMIFS('Monthly Detail'!$O$2:$O${last_row},'Monthly Detail'!$C$2:$C${last_row},A{flat_row},'Monthly Detail'!$B$2:$B${last_row},B{flat_row})").font = FORMULA_FONT
        ws3.cell(row=flat_row, column=5, value=f"=D{flat_row}-C{flat_row}").font = FORMULA_FONT
        ws3.cell(row=flat_row, column=6, value=f"=IFERROR(E{flat_row}/C{flat_row},0)").font = FORMULA_FONT
        for col in (3, 4, 5):
            ws3.cell(row=flat_row, column=col).number_format = MONEY
        ws3.cell(row=flat_row, column=6).number_format = PCT1
        combo_rows[(cat, reg)] = flat_row
        flat_row += 1
flat_last_row = flat_row - 1

for c, w in zip(range(1, 7), [22, 12, 14, 14, 12, 11]):
    ws3.column_dimensions[get_column_letter(c)].width = w

# pivoted heatmap matrix
matrix_top = flat_last_row + 3
ws3.cell(row=matrix_top, column=1, value="Variance $ Matrix").font = Font(name=FONT, bold=True, size=12)
header_row = matrix_top + 1
ws3.cell(row=header_row, column=1, value="Category").font = HEAD_FONT
ws3.cell(row=header_row, column=1).fill = HEAD_FILL
for j, reg in enumerate(regions, start=2):
    cell = ws3.cell(row=header_row, column=j, value=reg)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
for i, cat in enumerate(categories, start=header_row + 1):
    ws3.cell(row=i, column=1, value=cat).font = FORMULA_FONT
    for j, reg in enumerate(regions, start=2):
        r = combo_rows[(cat, reg)]
        cell = ws3.cell(row=i, column=j, value=f"=E{r}")
        cell.font = FORMULA_FONT
        cell.number_format = MONEY
        cell.border = BORDER
matrix_bottom = header_row + len(categories)

heat_range = f"B{header_row+1}:E{matrix_bottom}"
ws3.conditional_formatting.add(
    heat_range,
    ColorScaleRule(
        start_type="min", start_color="C0392B",
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="max", end_color="1E8449",
    ),
)

# =========================================================== Margin Bridge
ws4 = wb.create_sheet("Margin Bridge")
ws4.sheet_view.showGridLines = False
ws4["A1"] = "Margin Bridge — Budget Profit to Actual Profit (24 months, network-wide)"
ws4["A1"].font = Font(name=FONT, bold=True, size=13)
ws4["A2"] = "Each effect isolates one lever, holding the others at budget/actual as noted — the three effects sum exactly to total variance."
ws4["A2"].font = SUB_FONT

bridge_headers = ["Step", "Amount"]
for c, h in enumerate(bridge_headers, start=1):
    cell = ws4.cell(row=4, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL

bridge_rows = [
    ("Budget Profit", f"=SUM('Monthly Detail'!I2:I{last_row})"),
    ("Volume/Revenue Effect", f"=SUM('Monthly Detail'!Q2:Q{last_row})"),
    ("Discount Effect", f"=SUM('Monthly Detail'!R2:R{last_row})"),
    ("COGS Effect", f"=SUM('Monthly Detail'!S2:S{last_row})"),
    ("Actual Profit", f"=B5+B6+B7+B8"),
]
for i, (label, formula) in enumerate(bridge_rows, start=5):
    ws4.cell(row=i, column=1, value=label).font = FORMULA_FONT
    cell = ws4.cell(row=i, column=2, value=formula)
    cell.font = Font(name=FONT, bold=(label in ("Budget Profit", "Actual Profit")))
    cell.number_format = MONEY
ws4.column_dimensions["A"].width = 24
ws4.column_dimensions["B"].width = 16

chart = BarChart()
chart.type = "col"
chart.title = "Margin Bridge: Budget → Actual Profit"
chart.y_axis.title = "USD"
chart.x_axis.title = None
data = Reference(ws4, min_col=2, min_row=4, max_row=9)
cats = Reference(ws4, min_col=1, min_row=5, max_row=9)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width, chart.height = 18, 10
ws4.add_chart(chart, "D4")

# =========================================================== Top Variance Drivers
ws5 = wb.create_sheet("Top Variance Drivers")
ws5.sheet_view.showGridLines = False
ws5["A1"] = "Top 8 Variance Drivers (most negative first)"
ws5["A1"].font = Font(name=FONT, bold=True, size=13)
ws5["A2"] = "Ranked with SMALL() over a tie-broken key, pulled back with INDEX/MATCH — no manual sorting."
ws5["A2"].font = SUB_FONT

# helper column on the summary sheet: tie-broken key so SMALL() never collides
ws3.cell(row=4, column=8, value="Tiebreak Key").font = HEAD_FONT
ws3.cell(row=4, column=8).fill = HEAD_FILL
for cat, reg in combo_rows:
    r = combo_rows[(cat, reg)]
    ws3.cell(row=r, column=8, value=f"=E{r}+ROW()/1000000").font = FORMULA_FONT

rank_headers = ["Rank", "Category", "Region", "Variance $", "Variance %"]
for c, h in enumerate(rank_headers, start=1):
    cell = ws5.cell(row=4, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL

for k in range(1, 9):
    r = 4 + k
    ws5.cell(row=r, column=1, value=k).font = FORMULA_FONT
    key_formula = f"SMALL('Region x Category Summary'!$H$5:$H${flat_last_row},{k})"
    match_formula = f"MATCH({key_formula},'Region x Category Summary'!$H$5:$H${flat_last_row},0)"
    ws5.cell(row=r, column=2, value=f"=INDEX('Region x Category Summary'!$A$5:$A${flat_last_row},{match_formula})").font = FORMULA_FONT
    ws5.cell(row=r, column=3, value=f"=INDEX('Region x Category Summary'!$B$5:$B${flat_last_row},{match_formula})").font = FORMULA_FONT
    ws5.cell(row=r, column=4, value=f"=INDEX('Region x Category Summary'!$E$5:$E${flat_last_row},{match_formula})").font = FORMULA_FONT
    ws5.cell(row=r, column=4).number_format = MONEY
    ws5.cell(row=r, column=5, value=f"=INDEX('Region x Category Summary'!$F$5:$F${flat_last_row},{match_formula})").font = FORMULA_FONT
    ws5.cell(row=r, column=5).number_format = PCT1
for c, w in zip(range(1, 6), [7, 22, 12, 13, 11]):
    ws5.column_dimensions[get_column_letter(c)].width = w

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Workbook written to {OUT}")
