"""
Extra 07 -- Interactive budget workbook.

A 24-month budget model with a live scenario toggle (data-validation
dropdown + INDEX/MATCH driving every downstream formula) and a side-by-side
scenario comparison sheet that recomputes all three scenarios independently
of the toggle, using closed-form compound-growth formulas so it doesn't
need three duplicate 24-month grids.
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import LineChart, Reference

BASE = Path(__file__).resolve().parents[1]
OUT = BASE / "excel" / "budget_scenario_model.xlsx"
OUT.parent.mkdir(parents=True, exist_ok=True)

FONT = "Arial"
BLUE = Font(name=FONT, color="0000FF")
BLUE_BOLD = Font(name=FONT, color="0000FF", bold=True)
BLACK = Font(name=FONT, color="000000")
BLACK_BOLD = Font(name=FONT, color="000000", bold=True)
GREEN = Font(name=FONT, color="008000")
HEADER_FONT = Font(name=FONT, bold=True, size=13)
SECTION_FONT = Font(name=FONT, bold=True, size=11, color="44603B")
YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="E4EBE1")
THIN = Side(style="thin", color="B9C2C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CUR = '$#,##0;($#,##0);"-"'
PCT = '0.0%;(0.0%);"-"'

wb = Workbook()

# ================================================================== README
ws = wb.active
ws.title = "README"
ws.column_dimensions["A"].width = 100
ws["A1"] = "Interactive Budget Workbook — 24-Month Scenario Model"
ws["A1"].font = HEADER_FONT
lines = [
    "",
    "This workbook models a 24-month (2026-2027) budget with a live scenario toggle.",
    "Change the dropdown on the Assumptions tab and every number on Budget Model recalculates —",
    "no copy-pasting between scenarios.",
    "",
    "How to use it:",
    "  1. Go to Assumptions and pick a scenario from the dropdown in cell B3.",
    "  2. Budget Model shows the full 24-month P&L for whichever scenario is selected.",
    "  3. Scenario Comparison shows all three scenarios side by side at once, independent",
    "     of the toggle, so you don't have to flip through one at a time to compare them.",
    "",
    "Color legend:",
    "  Blue text        an input you can change (assumptions, scenario drivers)",
    "  Yellow fill       the scenario selector, and other key assumptions worth double-checking",
    "  Black text        a formula on the same sheet",
    "  Green text        a formula that pulls a value from another sheet",
    "",
    "Sheets:",
    "  Assumptions           shared assumptions, per-scenario drivers, and the scenario toggle",
    "  Budget Model           24-month P&L for the currently selected scenario",
    "  Scenario Comparison    all three scenarios side by side, with a chart",
]
for i, line in enumerate(lines, start=2):
    ws[f"A{i}"] = line
    ws[f"A{i}"].font = BLACK_BOLD if line and not line.startswith(" ") and ":" in line and i > 6 else BLACK

# ================================================================== Assumptions
ws = wb.create_sheet("Assumptions")
ws.column_dimensions["A"].width = 32
for col in "BCDE":
    ws.column_dimensions[col].width = 20

ws["A1"] = "Budget Assumptions"
ws["A1"].font = HEADER_FONT

ws["A3"] = "Scenario:"
ws["A3"].font = BLACK_BOLD
ws["B3"] = "Base"
ws["B3"].font = BLACK_BOLD
ws["B3"].fill = YELLOW_FILL
ws["B3"].border = BORDER
ws["B3"].alignment = Alignment(horizontal="center")
dv = DataValidation(type="list", formula1='"Conservative,Base,Aggressive"', allow_blank=False)
dv.error = "Pick one of: Conservative, Base, Aggressive"
dv.prompt = "Choose a scenario"
ws.add_data_validation(dv)
dv.add(ws["B3"])

ws["A5"] = "Shared Assumptions (all scenarios)"
ws["A5"].font = SECTION_FONT
shared = [
    ("Starting Monthly Revenue ($)", 400000, CUR),
    ("Starting Headcount", 18, "0"),
    ("Fixed Opex per Month ($)", 45000, CUR),
]
for i, (label, val, fmt) in enumerate(shared, start=6):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = BLACK
    ws[f"B{i}"] = val
    ws[f"B{i}"].font = BLUE
    ws[f"B{i}"].number_format = fmt
    ws[f"B{i}"].border = BORDER
# named single cells for readability below
START_REV, START_HC, FIXED_OPEX = "B6", "B7", "B8"

ws["A10"] = "Scenario Definitions"
ws["A10"].font = SECTION_FONT
headers = ["Scenario", "Monthly Growth Rate", "Gross Margin %", "Headcount Adds / Qtr", "Cost per Head / Month ($)"]
for j, h in enumerate(headers):
    c = ws.cell(row=11, column=1 + j, value=h)
    c.font = BLACK_BOLD
    c.fill = HEADER_FILL
    c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="center")

scenarios = [
    ("Conservative", 0.010, 0.62, 0, 9500),
    ("Base",         0.025, 0.65, 1, 9500),
    ("Aggressive",   0.045, 0.68, 3, 9500),
]
for i, (name, growth, margin, hc_add, cost) in enumerate(scenarios, start=12):
    ws.cell(row=i, column=1, value=name).font = BLUE
    ws.cell(row=i, column=2, value=growth).font = BLUE
    ws.cell(row=i, column=2).number_format = PCT
    ws.cell(row=i, column=3, value=margin).font = BLUE
    ws.cell(row=i, column=3).number_format = PCT
    ws.cell(row=i, column=4, value=hc_add).font = BLUE
    ws.cell(row=i, column=5, value=cost).font = BLUE
    ws.cell(row=i, column=5).number_format = CUR
    for col in range(1, 6):
        ws.cell(row=i, column=col).border = BORDER
# scenario rows are 12 (Conservative), 13 (Base), 14 (Aggressive)

ws["A16"] = "Active Scenario Drivers (auto-updates from B3)"
ws["A16"].font = SECTION_FONT
active_rows = [
    ("Monthly Growth Rate", "B12:B14", PCT),
    ("Gross Margin %", "C12:C14", PCT),
    ("Headcount Adds / Qtr", "D12:D14", "0"),
    ("Cost per Head / Month ($)", "E12:E14", CUR),
]
for i, (label, rng, fmt) in enumerate(active_rows, start=17):
    ws[f"A{i}"] = label
    ws[f"A{i}"].font = BLACK
    ws[f"B{i}"] = f"=INDEX({rng},MATCH($B$3,$A$12:$A$14,0))"
    ws[f"B{i}"].font = BLACK_BOLD
    ws[f"B{i}"].number_format = fmt
    ws[f"B{i}"].border = BORDER
# Active drivers: growth=B17, margin=B18, hc_add=B19, cost_per_head=B20

wb.save(OUT)
print("Assumptions sheet written")

# ================================================================== Budget Model
ws = wb.create_sheet("Budget Model")
ws.column_dimensions["A"].width = 26
N_MONTHS = 24
for m in range(1, N_MONTHS + 1):
    ws.column_dimensions[get_column_letter(1 + m)].width = 11

ws["A1"] = "Budget Model — Active Scenario"
ws["A1"].font = HEADER_FONT
ws["A2"] = "Active Scenario:"
ws["A2"].font = BLACK_BOLD
ws["B2"] = "=Assumptions!$B$3"
ws["B2"].font = GREEN
ws["B2"].font = Font(name=FONT, color="008000", bold=True)

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
ROW_MONTH_LABEL, ROW_MONTH_NUM = 4, 5
ROW_HC, ROW_REV, ROW_COGS, ROW_GP, ROW_OPEX, ROW_EBITDA, ROW_MARGIN = 6, 7, 8, 9, 10, 11, 12

ws.cell(row=ROW_MONTH_LABEL, column=1, value="Month").font = BLACK_BOLD
ws.cell(row=ROW_MONTH_NUM, column=1, value="Month #").font = BLACK
ws.cell(row=ROW_HC, column=1, value="Headcount").font = BLACK
ws.cell(row=ROW_REV, column=1, value="Revenue ($)").font = BLACK_BOLD
ws.cell(row=ROW_COGS, column=1, value="COGS ($)").font = BLACK
ws.cell(row=ROW_GP, column=1, value="Gross Profit ($)").font = BLACK
ws.cell(row=ROW_OPEX, column=1, value="Opex ($)").font = BLACK
ws.cell(row=ROW_EBITDA, column=1, value="EBITDA ($)").font = BLACK_BOLD
ws.cell(row=ROW_MARGIN, column=1, value="EBITDA Margin %").font = BLACK

for m in range(1, N_MONTHS + 1):
    col = 1 + m
    L = get_column_letter(col)
    Lprev = get_column_letter(col - 1)
    year_offset = 26 if m > 12 else 25
    month_name = MONTH_NAMES[(m - 1) % 12]
    year_label = "26" if m <= 12 else "27"

    ws.cell(row=ROW_MONTH_LABEL, column=col, value=f"{month_name}-{year_label}").font = BLACK_BOLD
    ws.cell(row=ROW_MONTH_LABEL, column=col).alignment = Alignment(horizontal="center")
    ws.cell(row=ROW_MONTH_NUM, column=col, value=m).font = BLACK
    ws.cell(row=ROW_MONTH_NUM, column=col).alignment = Alignment(horizontal="center")

    # headcount: starting HC + (adds/qtr) x (completed quarters so far)
    hc_formula = f"=Assumptions!$B$7+Assumptions!$B$19*(ROUNDUP({L}${ROW_MONTH_NUM}/3,0)-1)"
    c = ws.cell(row=ROW_HC, column=col, value=hc_formula)
    c.font = GREEN
    c.number_format = "0"

    # revenue: first month = starting revenue; every other month = prior month x (1+growth)
    if m == 1:
        rev_formula = "=Assumptions!$B$6"
    else:
        rev_formula = f"={Lprev}${ROW_REV}*(1+Assumptions!$B$17)"
    c = ws.cell(row=ROW_REV, column=col, value=rev_formula)
    c.font = GREEN
    c.number_format = CUR

    cogs_formula = f"={L}${ROW_REV}*(1-Assumptions!$B$18)"
    c = ws.cell(row=ROW_COGS, column=col, value=cogs_formula)
    c.font = GREEN
    c.number_format = CUR

    gp_formula = f"={L}${ROW_REV}-{L}${ROW_COGS}"
    c = ws.cell(row=ROW_GP, column=col, value=gp_formula)
    c.font = BLACK
    c.number_format = CUR

    opex_formula = f"=Assumptions!$B$8+{L}${ROW_HC}*Assumptions!$B$20"
    c = ws.cell(row=ROW_OPEX, column=col, value=opex_formula)
    c.font = GREEN
    c.number_format = CUR

    ebitda_formula = f"={L}${ROW_GP}-{L}${ROW_OPEX}"
    c = ws.cell(row=ROW_EBITDA, column=col, value=ebitda_formula)
    c.font = BLACK_BOLD
    c.number_format = CUR

    margin_formula = f"=IFERROR({L}${ROW_EBITDA}/{L}${ROW_REV},0)"
    c = ws.cell(row=ROW_MARGIN, column=col, value=margin_formula)
    c.font = BLACK
    c.number_format = PCT

# annual summary
ROW_SUMMARY_HDR = ROW_MARGIN + 2
ws.cell(row=ROW_SUMMARY_HDR, column=1, value="Annual Summary").font = SECTION_FONT
ws.cell(row=ROW_SUMMARY_HDR + 1, column=2, value="Year 1 (2026)").font = BLACK_BOLD
ws.cell(row=ROW_SUMMARY_HDR + 1, column=3, value="Year 2 (2027)").font = BLACK_BOLD
for col in (2, 3):
    ws.cell(row=ROW_SUMMARY_HDR + 1, column=col).alignment = Alignment(horizontal="center")

Y1_START, Y1_END = get_column_letter(2), get_column_letter(13)   # months 1-12 -> cols B..M
Y2_START, Y2_END = get_column_letter(14), get_column_letter(25)  # months 13-24 -> cols N..Y

summary_rows = [
    ("Revenue ($)", ROW_REV, CUR, "SUM"),
    ("EBITDA ($)", ROW_EBITDA, CUR, "SUM"),
]
r = ROW_SUMMARY_HDR + 2
for label, src_row, fmt, agg in summary_rows:
    ws.cell(row=r, column=1, value=label).font = BLACK
    ws.cell(row=r, column=2, value=f"=SUM({Y1_START}{src_row}:{Y1_END}{src_row})").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = fmt
    ws.cell(row=r, column=3, value=f"=SUM({Y2_START}{src_row}:{Y2_END}{src_row})").font = BLACK_BOLD
    ws.cell(row=r, column=3).number_format = fmt
    r += 1
ws.cell(row=r, column=1, value="EBITDA Margin %").font = BLACK
ws.cell(row=r, column=2, value=f"=IFERROR(B{r-1}/B{r-2},0)").font = BLACK_BOLD
ws.cell(row=r, column=2).number_format = PCT
ws.cell(row=r, column=3, value=f"=IFERROR(C{r-1}/C{r-2},0)").font = BLACK_BOLD
ws.cell(row=r, column=3).number_format = PCT

ws.freeze_panes = "B6"
wb.save(OUT)
print("Budget Model sheet written")

# ================================================================== Scenario Comparison
ws = wb.create_sheet("Scenario Comparison")
ws.column_dimensions["A"].width = 12
for col in "BCDEFGHIJ":
    ws.column_dimensions[col].width = 16

ws["A1"] = "Scenario Comparison — quarterly, all three scenarios at once"
ws["A1"].font = HEADER_FONT
ws["A2"] = "Independent of the Assumptions toggle -- each column pulls its own scenario row directly."
ws["A2"].font = Font(name=FONT, italic=True, size=9, color="898781")

SCEN_ROWS = {"Conservative": 12, "Base": 13, "Aggressive": 14}  # rows on Assumptions sheet
N_Q = 8

def q_col(scen_index, metric_index):
    # layout: Quarter | Con Rev | Base Rev | Agg Rev | Con EBITDA | Base EBITDA | Agg EBITDA
    return 2 + metric_index * 3 + scen_index

ws.cell(row=4, column=1, value="Quarter").font = BLACK_BOLD
metric_labels = ["Revenue ($)", "EBITDA ($)"]
scen_names = list(SCEN_ROWS.keys())
for metric_index, mlabel in enumerate(metric_labels):
    for scen_index, sname in enumerate(scen_names):
        col = q_col(scen_index, metric_index)
        c = ws.cell(row=3, column=col, value=mlabel if scen_index == 1 else None)
        c.font = BLACK_BOLD
        c2 = ws.cell(row=4, column=col, value=sname)
        c2.font = BLACK_BOLD
        c2.fill = HEADER_FILL
        c2.border = BORDER
        c2.alignment = Alignment(horizontal="center")

for q in range(1, N_Q + 1):
    row = 4 + q
    month_end = q * 3
    ws.cell(row=row, column=1, value=f"Q{q}").font = BLACK

    for scen_index, sname in enumerate(scen_names):
        srow = SCEN_ROWS[sname]
        growth_ref = f"Assumptions!$B${srow}"
        margin_ref = f"Assumptions!$C${srow}"
        hc_add_ref = f"Assumptions!$D${srow}"
        # revenue at end of this quarter: closed-form compound growth from the starting revenue.
        # Budget Model treats month 1 as the starting revenue itself (growth applies from month 2
        # onward), so month N has had (N-1) compounding steps -- match that here exactly, or this
        # sheet and Budget Model quietly disagree by one period.
        rev_formula = f"=Assumptions!$B$6*(1+{growth_ref})^({month_end}-1)"
        rev_col = q_col(scen_index, 0)
        c = ws.cell(row=row, column=rev_col, value=rev_formula)
        c.font = GREEN
        c.number_format = CUR

        # headcount and opex at end of this quarter, then EBITDA
        rev_cell = f"{get_column_letter(rev_col)}{row}"
        ebitda_formula = (
            f"={rev_cell}*{margin_ref}"
            f"-(Assumptions!$B$8+(Assumptions!$B$7+{hc_add_ref}*({q}-1))*Assumptions!$E${srow})"
        )
        ebitda_col = q_col(scen_index, 1)
        c = ws.cell(row=row, column=ebitda_col, value=ebitda_formula)
        c.font = GREEN
        c.number_format = CUR

# summary block: Q8 exit values + margin, per scenario
ROW_SUM_HDR = 4 + N_Q + 2
ws.cell(row=ROW_SUM_HDR, column=1, value="Year 2 Exit (Q8) Summary").font = SECTION_FONT
ws.cell(row=ROW_SUM_HDR + 1, column=1, value="Scenario").font = BLACK_BOLD
ws.cell(row=ROW_SUM_HDR + 1, column=2, value="Revenue ($)").font = BLACK_BOLD
ws.cell(row=ROW_SUM_HDR + 1, column=3, value="EBITDA ($)").font = BLACK_BOLD
ws.cell(row=ROW_SUM_HDR + 1, column=4, value="EBITDA Margin %").font = BLACK_BOLD
q8_row = 4 + N_Q
for scen_index, sname in enumerate(scen_names):
    r = ROW_SUM_HDR + 2 + scen_index
    ws.cell(row=r, column=1, value=sname).font = BLACK
    rev_cell = f"{get_column_letter(q_col(scen_index, 0))}{q8_row}"
    ebitda_cell = f"{get_column_letter(q_col(scen_index, 1))}{q8_row}"
    ws.cell(row=r, column=2, value=f"={rev_cell}").font = BLACK_BOLD
    ws.cell(row=r, column=2).number_format = CUR
    ws.cell(row=r, column=3, value=f"={ebitda_cell}").font = BLACK_BOLD
    ws.cell(row=r, column=3).number_format = CUR
    ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)").font = BLACK_BOLD
    ws.cell(row=r, column=4).number_format = PCT

# chart: quarterly revenue, all three scenarios
chart = LineChart()
chart.title = "Quarterly Revenue by Scenario"
chart.style = 2
chart.y_axis.title = "Revenue ($)"
chart.x_axis.title = "Quarter"
chart.height = 8
chart.width = 18
cats = Reference(ws, min_col=1, min_row=5, max_row=4 + N_Q)
for scen_index, sname in enumerate(scen_names):
    rev_col = q_col(scen_index, 0)
    data = Reference(ws, min_col=rev_col, min_row=4, max_row=4 + N_Q)
    chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, f"A{ROW_SUM_HDR + 7}")

wb.save(OUT)
print("Scenario Comparison sheet written")
print(f"\nSaved to {OUT}")
