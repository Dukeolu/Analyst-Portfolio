"""
Builds cycle_time_analysis.xlsx for BA Case 01 (Process Improvement).

Sheets:
  Overview            - legend, scope
  Invoice Sample       - 120 observed current-state invoices, formula-derived
                         tier/cycle-time/discount/late-fee columns
  Bottleneck Analysis  - AVERAGEIFS/COUNTIFS/SUMPRODUCT rollups from the sample
  Future-State Model   - assumption inputs (blue/yellow) + row-level future
                         projection mirroring the sample + annualized $ impact
  Summary              - headline KPIs, all formula-linked
"""
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

BASE = Path(__file__).resolve().parents[1]
DATA = BASE / "data" / "invoice_sample.csv"
OUT = BASE / "artifacts" / "cycle_time_analysis.xlsx"
OUT.parent.mkdir(exist_ok=True)

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="44603B")
YELLOW = PatternFill("solid", fgColor="FFFF00")
LIGHT = PatternFill("solid", fgColor="F3F5F4")
THIN = Side(style="thin", color="B9C2C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="44603B")

wb = Workbook()

def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ---------------------------------------------------------------- Overview
ws = wb.active
ws.title = "Overview"
ws["B2"] = "Case 01 — AP Invoice Cycle-Time Analysis"
ws["B2"].font = TITLE_FONT
ws["B4"] = ("Supports the Process Improvement Charter. Brightpath Distribution's AP "
            "invoice approval process is flat: every invoice $500+ requires the same "
            "three sequential approvals regardless of size. This workbook measures the "
            "current-state cost of that in one month of observed data, then models a "
            "tiered, tool-assisted future state.")
ws["B4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 48

ws["B6"] = "Color legend"
ws["B6"].font = BOLD
legend = [
    ("Blue text", "Hardcoded input / assumption you can change", BLUE),
    ("Yellow fill", "Key assumption — worth double-checking before relying on the output", None),
    ("Black text", "Formula — recalculates automatically", BLACK),
    ("Green text", "Link pulled from another sheet", GREEN),
]
r = 7
for label, desc, font in legend:
    ws.cell(row=r, column=2, value=label).font = font or BOLD
    if font is None:
        ws.cell(row=r, column=2).fill = YELLOW
    ws.cell(row=r, column=3, value=desc)
    r += 1

ws["B12"] = "Sheets"
ws["B12"].font = BOLD
sheets_desc = [
    "Invoice Sample — 120 observed invoices from one month, current-state process",
    "Bottleneck Analysis — where the time actually goes, by stage and by tier",
    "Future-State Model — tiered approvals + routing tool, projected row by row",
    "Summary — headline before/after KPIs",
]
r = 13
for s in sheets_desc:
    ws.cell(row=r, column=2, value=f"• {s}")
    r += 1
autosize(ws, [3, 14, 70])

# ---------------------------------------------------------------- Invoice Sample
ws2 = wb.create_sheet("Invoice Sample")
headers = ["Invoice ID", "Vendor", "Vendor Type", "Department", "Amount", "Payment Terms",
           "Days: Data Entry", "Days: Dept Mgr Approval", "Days: Finance Mgr Approval",
           "Days: Controller Approval", "Days: Payment Processing", "Total Cycle Days",
           "Tier", "Discount Eligible?", "Discount Captured?", "Late Fee Incurred?",
           "Discount $ Missed", "Late Fee $ Incurred"]
for c, h in enumerate(headers, start=1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(headers))
ws2.freeze_panes = "A2"

with open(DATA) as f:
    rdr = list(csv.DictReader(f))

for i, row in enumerate(rdr):
    r = i + 2
    ws2.cell(row=r, column=1, value=row["Invoice ID"]).font = BLUE
    ws2.cell(row=r, column=2, value=row["Vendor"]).font = BLUE
    ws2.cell(row=r, column=3, value=row["Vendor Type"]).font = BLUE
    ws2.cell(row=r, column=4, value=row["Department"]).font = BLUE
    ws2.cell(row=r, column=5, value=float(row["Amount"])).font = BLUE
    ws2.cell(row=r, column=5).number_format = "$#,##0.00"
    ws2.cell(row=r, column=6, value=row["Payment Terms"]).font = BLUE
    ws2.cell(row=r, column=7, value=float(row["Days: Data Entry"])).font = BLUE
    ws2.cell(row=r, column=8, value=float(row["Days: Dept Mgr Approval"])).font = BLUE
    ws2.cell(row=r, column=9, value=float(row["Days: Finance Mgr Approval"])).font = BLUE
    ws2.cell(row=r, column=10, value=float(row["Days: Controller Approval"])).font = BLUE
    ws2.cell(row=r, column=11, value=float(row["Days: Payment Processing"])).font = BLUE

    # formula columns
    ws2.cell(row=r, column=12, value=f"=SUM(G{r}:K{r})").font = BLACK
    ws2.cell(row=r, column=12).number_format = "0.0"
    ws2.cell(row=r, column=13,
             value=f'=IF(E{r}<500,"Tier 0 (<$500)",IF(E{r}<10000,"Tier 1 ($500-$10K)",'
                   f'IF(E{r}<50000,"Tier 2 ($10K-$50K)","Tier 3 (>$50K)")))').font = BLACK
    ws2.cell(row=r, column=14, value=f'=IF(F{r}="2/10 Net 30","Yes","No")').font = BLACK
    ws2.cell(row=r, column=15, value=f'=IF(AND(N{r}="Yes",(G{r}+H{r}+I{r}+J{r})<=10),"Yes","No")').font = BLACK
    ws2.cell(row=r, column=16, value=f'=IF(L{r}>30,"Yes","No")').font = BLACK
    ws2.cell(row=r, column=17, value=f'=IF(AND(N{r}="Yes",O{r}="No"),E{r}*0.02,0)').font = BLACK
    ws2.cell(row=r, column=17).number_format = "$#,##0.00"
    ws2.cell(row=r, column=18, value=f'=IF(P{r}="Yes",E{r}*0.015,0)').font = BLACK
    ws2.cell(row=r, column=18).number_format = "$#,##0.00"

    for c in range(1, 19):
        ws2.cell(row=r, column=c).border = BORDER

autosize(ws2, [11, 26, 12, 15, 12, 14, 11, 11, 11, 11, 11, 11, 17, 12, 12, 12, 13, 13])
LASTROW = len(rdr) + 1

# ---------------------------------------------------------------- Bottleneck Analysis
ws3 = wb.create_sheet("Bottleneck Analysis")
ws3["B2"] = "Bottleneck Analysis — current state"
ws3["B2"].font = TITLE_FONT
ws3.merge_cells("B2:F2")

ws3["B4"] = "Average days by stage (all invoices unless noted)"
ws3["B4"].font = BOLD
stage_rows = [
    ("Data entry — recurring vendors", f'=AVERAGEIFS(\'Invoice Sample\'!G2:G{LASTROW},\'Invoice Sample\'!C2:C{LASTROW},"Recurring")'),
    ("Data entry — one-off vendors", f'=AVERAGEIFS(\'Invoice Sample\'!G2:G{LASTROW},\'Invoice Sample\'!C2:C{LASTROW},"One-off")'),
    ("Dept Mgr approval (all invoices)", f'=AVERAGE(\'Invoice Sample\'!H2:H{LASTROW})'),
    ("Finance Mgr approval ($500+ only)", f'=AVERAGEIF(\'Invoice Sample\'!E2:E{LASTROW},">=500",\'Invoice Sample\'!I2:I{LASTROW})'),
    ("Controller approval ($500+ only)", f'=AVERAGEIF(\'Invoice Sample\'!E2:E{LASTROW},">=500",\'Invoice Sample\'!J2:J{LASTROW})'),
    ("Payment processing", f'=AVERAGE(\'Invoice Sample\'!K2:K{LASTROW})'),
]
r = 5
for label, formula in stage_rows:
    ws3.cell(row=r, column=2, value=label)
    ws3.cell(row=r, column=5, value=formula).font = GREEN
    ws3.cell(row=r, column=5).number_format = "0.00"
    r += 1

ws3["B13"] = "Invoice mix by tier (this sample)"
ws3["B13"].font = BOLD
ws3.append([])
tier_hdr_row = 14
for c, h in enumerate(["Tier", "Count", "% of invoices", "Avg amount", "Avg total cycle days"], start=2):
    ws3.cell(row=tier_hdr_row, column=c, value=h)
style_header_row(ws3, tier_hdr_row, 0)
for c in range(2, 7):
    ws3.cell(row=tier_hdr_row, column=c).font = HDR_FONT
    ws3.cell(row=tier_hdr_row, column=c).fill = HDR_FILL

tiers = ["Tier 0 (<$500)", "Tier 1 ($500-$10K)", "Tier 2 ($10K-$50K)", "Tier 3 (>$50K)"]
r = tier_hdr_row + 1
for t in tiers:
    ws3.cell(row=r, column=2, value=t)
    ws3.cell(row=r, column=3, value=f'=COUNTIF(\'Invoice Sample\'!M2:M{LASTROW},B{r})').font = GREEN
    ws3.cell(row=r, column=4, value=f'=C{r}/{LASTROW-1}').font = BLACK
    ws3.cell(row=r, column=4).number_format = "0.0%"
    ws3.cell(row=r, column=5, value=f'=AVERAGEIF(\'Invoice Sample\'!M2:M{LASTROW},B{r},\'Invoice Sample\'!E2:E{LASTROW})').font = GREEN
    ws3.cell(row=r, column=5).number_format = "$#,##0"
    ws3.cell(row=r, column=6, value=f'=AVERAGEIF(\'Invoice Sample\'!M2:M{LASTROW},B{r},\'Invoice Sample\'!L2:L{LASTROW})').font = GREEN
    ws3.cell(row=r, column=6).number_format = "0.0"
    for c in range(2, 7):
        ws3.cell(row=r, column=c).border = BORDER
    r += 1
TIER_ROW0 = tier_hdr_row + 1

ws3["B21"] = "Current-state financial leakage (this sample, one month)"
ws3["B21"].font = BOLD
leak_rows = [
    ("Invoices with discount terms missing the 10-day window", f'=COUNTIFS(\'Invoice Sample\'!N2:N{LASTROW},"Yes",\'Invoice Sample\'!O2:O{LASTROW},"No")', "0"),
    ("Discount $ forfeited this month", f'=SUM(\'Invoice Sample\'!Q2:Q{LASTROW})', "$#,##0"),
    ("Invoices that missed Net 30 (late fee triggered)", f'=COUNTIF(\'Invoice Sample\'!P2:P{LASTROW},"Yes")', "0"),
    ("Late fee $ incurred this month", f'=SUM(\'Invoice Sample\'!R2:R{LASTROW})', "$#,##0"),
]
r = 22
for label, formula, fmt in leak_rows:
    ws3.cell(row=r, column=2, value=label)
    cell = ws3.cell(row=r, column=5, value=formula)
    cell.font = GREEN
    cell.number_format = fmt
    r += 1

autosize(ws3, [3, 34, 12, 14, 12, 14])

# ---------------------------------------------------------------- Future-State Model
ws4 = wb.create_sheet("Future-State Model")
ws4["B2"] = "Future-State Model — tiered approvals + routing tool"
ws4["B2"].font = TITLE_FONT
ws4.merge_cells("B2:H2")

ws4["B4"] = "Levers (change these to test other scenarios)"
ws4["B4"].font = BOLD
levers = [
    ("Required approvals — Tier 0 (<$500)", "Dept Mgr only (unchanged)", None),
    ("Required approvals — Tier 1 ($500-$10K)", "Dept Mgr only", None),
    ("Required approvals — Tier 2 ($10K-$50K)", "Dept Mgr + Finance Mgr", None),
    ("Required approvals — Tier 3 (>$50K)", "Dept Mgr + Finance Mgr + Controller (unchanged)", None),
]
r = 5
for label, val, _ in levers:
    ws4.cell(row=r, column=2, value=label)
    ws4.cell(row=r, column=6, value=val).font = BLUE
    ws4.cell(row=r, column=6).fill = YELLOW
    r += 1

r += 1
ROW_OCR = r
ws4.cell(row=r, column=2, value="OCR/auto-capture time reduction — recurring vendors")
ws4.cell(row=r, column=6, value=0.80).font = BLUE
ws4.cell(row=r, column=6).number_format = "0%"
ws4.cell(row=r, column=6).fill = YELLOW
r += 1
ROW_OCR_ONEOFF = r
ws4.cell(row=r, column=2, value="Data-entry time reduction — one-off / new vendors (not OCR-eligible)")
ws4.cell(row=r, column=6, value=0.0).font = BLUE
ws4.cell(row=r, column=6).number_format = "0%"
ws4.cell(row=r, column=6).fill = YELLOW
r += 1
ROW_APPROVAL_CUT = r
ws4.cell(row=r, column=2, value="Approval queue-time reduction — routing tool + defined backup approvers")
ws4.cell(row=r, column=6, value=0.42).font = BLUE
ws4.cell(row=r, column=6).number_format = "0%"
ws4.cell(row=r, column=6).fill = YELLOW
r += 1
ROW_PAY_CUT = r
ws4.cell(row=r, column=2, value="Payment-processing time reduction — batch scheduling automation")
ws4.cell(row=r, column=6, value=0.30).font = BLUE
ws4.cell(row=r, column=6).number_format = "0%"
ws4.cell(row=r, column=6).fill = YELLOW
r += 2

ws4.cell(row=r, column=2, value="Implementation cost (one-time + Year 1)").font = BOLD
r += 1
ROW_COST_TOOL = r
ws4.cell(row=r, column=2, value="Routing/workflow tool — annual subscription")
ws4.cell(row=r, column=6, value=18000).font = BLUE
ws4.cell(row=r, column=6).number_format = "$#,##0"
r += 1
ROW_COST_OCR = r
ws4.cell(row=r, column=2, value="OCR/invoice-capture tool — annual subscription")
ws4.cell(row=r, column=6, value=9600).font = BLUE
ws4.cell(row=r, column=6).number_format = "$#,##0"
r += 1
ROW_COST_IMPL = r
ws4.cell(row=r, column=2, value="One-time implementation & change-management effort")
ws4.cell(row=r, column=6, value=22000).font = BLUE
ws4.cell(row=r, column=6).number_format = "$#,##0"
r += 2

# per-stage projected durations, linked from Bottleneck Analysis
ws4.cell(row=r, column=2, value="Projected stage durations").font = BOLD
r += 1
HDR_PROJ = r
for c, h in enumerate(["Stage", "Current (obs.)", "Reduction", "Projected"], start=2):
    ws4.cell(row=r, column=c, value=h)
style_header_row(ws4, r, 0)
for c in range(2, 6):
    ws4.cell(row=r, column=c).font = HDR_FONT
    ws4.cell(row=r, column=c).fill = HDR_FILL
r += 1
ROW_DE_REC = r
ws4.cell(row=r, column=2, value="Data entry — recurring")
ws4.cell(row=r, column=3, value="='Bottleneck Analysis'!E5").font = GREEN
ws4.cell(row=r, column=4, value=f"=F{ROW_OCR}").font = GREEN
ws4.cell(row=r, column=4).number_format = "0%"
ws4.cell(row=r, column=5, value=f"=C{r}*(1-D{r})").font = BLACK
r += 1
ROW_DE_ONE = r
ws4.cell(row=r, column=2, value="Data entry — one-off")
ws4.cell(row=r, column=3, value="='Bottleneck Analysis'!E6").font = GREEN
ws4.cell(row=r, column=4, value=f"=F{ROW_OCR_ONEOFF}").font = GREEN
ws4.cell(row=r, column=4).number_format = "0%"
ws4.cell(row=r, column=5, value=f"=C{r}*(1-D{r})").font = BLACK
r += 1
ROW_DEPT = r
ws4.cell(row=r, column=2, value="Dept Mgr approval")
ws4.cell(row=r, column=3, value="='Bottleneck Analysis'!E7").font = GREEN
ws4.cell(row=r, column=4, value=f"=F{ROW_APPROVAL_CUT}").font = GREEN
ws4.cell(row=r, column=4).number_format = "0%"
ws4.cell(row=r, column=5, value=f"=C{r}*(1-D{r})").font = BLACK
r += 1
ROW_FIN = r
ws4.cell(row=r, column=2, value="Finance Mgr approval")
ws4.cell(row=r, column=3, value="='Bottleneck Analysis'!E8").font = GREEN
ws4.cell(row=r, column=4, value=f"=F{ROW_APPROVAL_CUT}").font = GREEN
ws4.cell(row=r, column=4).number_format = "0%"
ws4.cell(row=r, column=5, value=f"=C{r}*(1-D{r})").font = BLACK
r += 1
ROW_CTRL = r
ws4.cell(row=r, column=2, value="Controller approval")
ws4.cell(row=r, column=3, value="='Bottleneck Analysis'!E9").font = GREEN
ws4.cell(row=r, column=4, value=f"=F{ROW_APPROVAL_CUT}").font = GREEN
ws4.cell(row=r, column=4).number_format = "0%"
ws4.cell(row=r, column=5, value=f"=C{r}*(1-D{r})").font = BLACK
r += 1
ROW_PAY = r
ws4.cell(row=r, column=2, value="Payment processing")
ws4.cell(row=r, column=3, value="='Bottleneck Analysis'!E10").font = GREEN
ws4.cell(row=r, column=4, value=f"=F{ROW_PAY_CUT}").font = GREEN
ws4.cell(row=r, column=4).number_format = "0%"
ws4.cell(row=r, column=5, value=f"=C{r}*(1-D{r})").font = BLACK
for rr in range(HDR_PROJ + 1, ROW_PAY + 1):
    for c in range(2, 6):
        ws4.cell(row=rr, column=c).number_format = "0.00" if c != 4 else "0%"
        ws4.cell(row=rr, column=c).border = BORDER

autosize(ws4, [3, 48, 12, 12, 12, 30])

# --- row-level future projection, mirroring the sample
ws5 = wb.create_sheet("Future-State Projection")
headers5 = ["Invoice ID", "Vendor Type", "Amount", "Tier", "Payment Terms",
            "Projected Data Entry", "Projected Dept Mgr", "Projected Finance Mgr",
            "Projected Controller", "Projected Payment", "Projected Total Cycle Days",
            "Discount Eligible?", "Discount Captured (future)?", "Late Fee (future)?",
            "Discount $ Now Captured", "Late Fee $ Now Avoided"]
for c, h in enumerate(headers5, start=1):
    ws5.cell(row=1, column=c, value=h)
style_header_row(ws5, 1, len(headers5))
ws5.freeze_panes = "A2"

for i in range(len(rdr)):
    r = i + 2
    src = r  # same row number in Invoice Sample
    ws5.cell(row=r, column=1, value=f"='Invoice Sample'!A{src}").font = GREEN
    ws5.cell(row=r, column=2, value=f"='Invoice Sample'!C{src}").font = GREEN
    ws5.cell(row=r, column=3, value=f"='Invoice Sample'!E{src}").font = GREEN
    ws5.cell(row=r, column=3).number_format = "$#,##0.00"
    ws5.cell(row=r, column=4, value=f"='Invoice Sample'!M{src}").font = GREEN
    ws5.cell(row=r, column=5, value=f"='Invoice Sample'!F{src}").font = GREEN

    # projected data entry depends on vendor type
    ws5.cell(row=r, column=6,
             value=f"=IF(B{r}=\"Recurring\",'Future-State Model'!E{ROW_DE_REC},'Future-State Model'!E{ROW_DE_ONE})").font = BLACK

    # required approvals per tier (per the lever table): Tier0/1 = Dept only, Tier2 = Dept+Finance, Tier3 = all
    ws5.cell(row=r, column=7, value=f"='Future-State Model'!E{ROW_DEPT}").font = BLACK
    ws5.cell(row=r, column=8,
             value=f"=IF(OR(D{r}=\"Tier 2 ($10K-$50K)\",D{r}=\"Tier 3 (>$50K)\"),'Future-State Model'!E{ROW_FIN},0)").font = BLACK
    ws5.cell(row=r, column=9,
             value=f"=IF(D{r}=\"Tier 3 (>$50K)\",'Future-State Model'!E{ROW_CTRL},0)").font = BLACK
    ws5.cell(row=r, column=10, value=f"='Future-State Model'!E{ROW_PAY}").font = BLACK

    ws5.cell(row=r, column=11, value=f"=SUM(F{r}:J{r})").font = BLACK
    ws5.cell(row=r, column=11).number_format = "0.0"

    ws5.cell(row=r, column=12, value=f'=IF(E{r}="2/10 Net 30","Yes","No")').font = BLACK
    ws5.cell(row=r, column=13, value=f'=IF(AND(L{r}="Yes",(F{r}+G{r}+H{r}+I{r})<=10),"Yes","No")').font = BLACK
    ws5.cell(row=r, column=14, value=f'=IF(K{r}>30,"Yes","No")').font = BLACK

    # $ now captured = discount-eligible AND wasn't captured today (per Invoice Sample) AND is captured in future model
    ws5.cell(row=r, column=15,
             value=f'=IF(AND(L{r}="Yes",\'Invoice Sample\'!O{src}="No",M{r}="Yes"),C{r}*0.02,0)').font = BLACK
    ws5.cell(row=r, column=15).number_format = "$#,##0.00"
    ws5.cell(row=r, column=16,
             value=f'=IF(AND(\'Invoice Sample\'!P{src}="Yes",N{r}="No"),C{r}*0.015,0)').font = BLACK
    ws5.cell(row=r, column=16).number_format = "$#,##0.00"

    for c in range(1, 17):
        ws5.cell(row=r, column=c).border = BORDER

autosize(ws5, [11, 13, 12, 17, 13, 13, 12, 13, 13, 13, 15, 12, 15, 12, 14, 14])

# ---------------------------------------------------------------- Summary
ws6 = wb.create_sheet("Summary")
ws6["B2"] = "Summary — before / after"
ws6["B2"].font = TITLE_FONT
ws6.merge_cells("B2:F2")

ws6["B4"] = "Blended cycle time (population-weighted across this sample)"
ws6["B4"].font = BOLD
ws6["B5"] = "Current state — avg total cycle days"
ws6["E5"] = f"=AVERAGE('Invoice Sample'!L2:L{LASTROW})"
ws6["E5"].font = GREEN
ws6["E5"].number_format = "0.0"
ws6["B6"] = "Future state — avg total cycle days"
ws6["E6"] = f"=AVERAGE('Future-State Projection'!K2:K{LASTROW})"
ws6["E6"].font = GREEN
ws6["E6"].number_format = "0.0"
ws6["B7"] = "Reduction"
ws6["E7"] = "=(E5-E6)/E5"
ws6["E7"].font = BLACK
ws6["E7"].number_format = "0.0%"
ws6["B7"].font = BOLD
ws6["E7"].font = Font(name="Arial", size=10, bold=True)

ws6["B9"] = "Annualized financial impact (monthly sample × 12)"
ws6["B9"].font = BOLD
ws6["B10"] = "Months per year"
ws6["F10"] = 12
ws6["F10"].font = BLUE
ws6["F10"].fill = YELLOW

ws6["B11"] = "Discount $ captured / year (newly captured, future state)"
ws6["E11"] = f"=SUM('Future-State Projection'!O2:O{LASTROW})*F10"
ws6["E11"].font = GREEN
ws6["E11"].number_format = "$#,##0"

ws6["B12"] = "Late fee $ avoided / year"
ws6["E12"] = f"=SUM('Future-State Projection'!P2:P{LASTROW})*F10"
ws6["E12"].font = GREEN
ws6["E12"].number_format = "$#,##0"

ws6["B13"] = "Total annual benefit"
ws6["E13"] = "=E11+E12"
ws6["E13"].font = BOLD
ws6["B13"].font = BOLD
ws6["E13"].number_format = "$#,##0"

ws6["B15"] = "Implementation cost"
ws6["B15"].font = BOLD
ws6["B16"] = "Routing tool (annual)"
ws6["E16"] = f"='Future-State Model'!F{ROW_COST_TOOL}"
ws6["E16"].font = GREEN
ws6["E16"].number_format = "$#,##0"
ws6["B17"] = "OCR tool (annual)"
ws6["E17"] = f"='Future-State Model'!F{ROW_COST_OCR}"
ws6["E17"].font = GREEN
ws6["E17"].number_format = "$#,##0"
ws6["B18"] = "One-time implementation (Year 1 only)"
ws6["E18"] = f"='Future-State Model'!F{ROW_COST_IMPL}"
ws6["E18"].font = GREEN
ws6["E18"].number_format = "$#,##0"
ws6["B19"] = "Total Year 1 cost"
ws6["E19"] = "=E16+E17+E18"
ws6["B19"].font = BOLD
ws6["E19"].font = BOLD
ws6["E19"].number_format = "$#,##0"

ws6["B21"] = "Net benefit — Year 1"
ws6["E21"] = "=E13-E19"
ws6["B21"].font = BOLD
ws6["E21"].font = BOLD
ws6["E21"].number_format = "$#,##0"

ws6["B22"] = "Net benefit — Year 2+ (subscriptions only, no re-implementation)"
ws6["E22"] = "=E13-E16-E17"
ws6["B22"].font = BOLD
ws6["E22"].font = BOLD
ws6["E22"].number_format = "$#,##0"

for row in ws6.iter_rows(min_row=5, max_row=22, min_col=2, max_col=6):
    for cell in row:
        if cell.value is not None:
            cell.border = BORDER

autosize(ws6, [3, 46, 3, 3, 16])

wb.save(OUT)
print(f"Saved {OUT}")
