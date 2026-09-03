"""
Builds cost_benefit_analysis.xlsx for BA Case 04 (Business Case & Cost-Benefit).

Sheets:
  Overview             - legend, scope
  Assumptions           - all blue/yellow inputs
  Option B - Hire       - year-by-year headcount & cost to keep pace with volume by hiring
  Option C - Automate   - year-by-year headcount, tool cost, and error cost with OCR/RPA
  NPV & Payback         - net benefit of C vs B, discounted, payback period
  Summary               - headline KPIs
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "artifacts" / "cost_benefit_analysis.xlsx"
OUT.parent.mkdir(exist_ok=True)

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="44603B")
YELLOW = PatternFill("solid", fgColor="FFFF00")
WIN_FILL = PatternFill("solid", fgColor="E4EBE1")
THIN = Side(style="thin", color="B9C2C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="44603B")

wb = Workbook()

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c

# ---------------------------------------------------------------- Overview
ws = wb.active
ws.title = "Overview"
ws["B2"] = "Case 04 — Order-Entry Automation: Cost-Benefit Analysis"
ws["B2"].font = TITLE_FONT
ws["B4"] = ("Supports the Business Case. Brightpath Distribution's inbound sales-order entry team "
            "manually keys orders from fax, email PDF, and an EDI portal into the ERP. Order volume "
            "is growing 15%/year. This workbook compares two ways to keep pace over 3 years: keep "
            "hiring at current productivity (Option B), or invest in OCR/RPA automation (Option C).")
ws["B4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 56
ws["B6"] = "Color legend"
ws["B6"].font = BOLD
for i, (label, desc, font) in enumerate([
    ("Blue text", "Hardcoded input / assumption", BLUE),
    ("Yellow fill", "Key assumption", None),
    ("Black text", "Formula", BLACK),
    ("Green text", "Link from another sheet", GREEN),
], start=7):
    ws.cell(row=i, column=2, value=label).font = font or BOLD
    if font is None:
        ws.cell(row=i, column=2).fill = YELLOW
    ws.cell(row=i, column=3, value=desc)
autosize(ws, [3, 16, 74])

# ---------------------------------------------------------------- Assumptions
wsA = wb.create_sheet("Assumptions")
wsA["B2"] = "Assumptions"
wsA["B2"].font = TITLE_FONT
rows = [
    ("Current annual order volume", 66500, "0"),
    ("Volume growth rate / year", 0.15, "0%"),
    ("Current headcount (data-entry clerks)", 7, "0"),
    ("Orders processed / clerk / year (manual)", 9500, "0"),
    ("Fully loaded cost / clerk / year", 52000, "$#,##0"),
    ("One-time hiring & onboarding cost / new hire", 8000, "$#,##0"),
    ("Manual order error rate (needs correction/rework)", 0.025, "0.0%"),
    ("Automated (OCR-validated) order error rate", 0.006, "0.0%"),
    ("Average cost per order error (rework, reship, CS time)", 35, "$#,##0"),
    ("RPA/OCR platform subscription / year", 85000, "$#,##0"),
    ("One-time implementation cost (Year 1 only)", 120000, "$#,##0"),
    ("Ongoing template/maintenance cost / year", 15000, "$#,##0"),
    ("Automation adoption — Year 1 (% of volume auto-processed)", 0.65, "0%"),
    ("Automation adoption — Year 2", 0.75, "0%"),
    ("Automation adoption — Year 3", 0.80, "0%"),
    ("Discount rate (WACC)", 0.08, "0%"),
]
r = 4
CELL = {}
for label, val, fmt in rows:
    wsA.cell(row=r, column=2, value=label)
    c = wsA.cell(row=r, column=6, value=val)
    c.font = BLUE
    c.number_format = fmt
    c.fill = YELLOW
    CELL[label] = f"Assumptions!$F${r}"
    r += 1
autosize(wsA, [3, 52, 3, 3, 3, 16])

def A(key):
    return CELL[key]

# ---------------------------------------------------------------- Option B - Hire
wsB = wb.create_sheet("Option B - Hire")
wsB["B2"] = "Option B — Hire to Keep Pace"
wsB["B2"].font = TITLE_FONT
wsB.merge_cells("B2:H2")
hdr_row = 4
cols = ["Year", "Order volume", "Headcount needed", "New hires", "Labor cost", "Hiring cost", "Total cost"]
for c, h in enumerate(cols, start=2):
    hdr(wsB, hdr_row, c, h)

r = hdr_row + 1
prev_hc_cell = A("Current headcount (data-entry clerks)")
for yr in [1, 2, 3]:
    wsB.cell(row=r, column=2, value=yr)
    if yr == 1:
        vol_formula = f"={A('Current annual order volume')}*(1+{A('Volume growth rate / year')})"
    else:
        vol_formula = f"=C{r-1}*(1+{A('Volume growth rate / year')})"
    wsB.cell(row=r, column=3, value=vol_formula).font = BLACK
    wsB.cell(row=r, column=3).number_format = "0"
    wsB.cell(row=r, column=4, value=f"=CEILING(C{r}/{A('Orders processed / clerk / year (manual)')},1)").font = BLACK
    prev = prev_hc_cell if yr == 1 else f"D{r-1}"
    wsB.cell(row=r, column=5, value=f"=D{r}-{prev}").font = BLACK
    wsB.cell(row=r, column=6, value=f"=D{r}*{A('Fully loaded cost / clerk / year')}").font = BLACK
    wsB.cell(row=r, column=6).number_format = "$#,##0"
    wsB.cell(row=r, column=7, value=f"=E{r}*{A('One-time hiring & onboarding cost / new hire')}").font = BLACK
    wsB.cell(row=r, column=7).number_format = "$#,##0"
    wsB.cell(row=r, column=8, value=f"=F{r}+G{r}").font = BOLD
    wsB.cell(row=r, column=8).number_format = "$#,##0"
    for c in range(2, 9):
        wsB.cell(row=r, column=c).border = BORDER
    r += 1
B_ROWS = (hdr_row + 1, r - 1)
r += 1
wsB.cell(row=r, column=2, value="3-year total").font = BOLD
wsB.cell(row=r, column=8, value=f"=SUM(H{B_ROWS[0]}:H{B_ROWS[1]})").font = BOLD
wsB.cell(row=r, column=8).number_format = "$#,##0"
B_TOTAL_ROW = r
autosize(wsB, [3, 6, 13, 16, 12, 14, 13, 14])

# ---------------------------------------------------------------- Option C - Automate
wsC = wb.create_sheet("Option C - Automate")
wsC["B2"] = "Option C — Invest in OCR/RPA Automation"
wsC["B2"].font = TITLE_FONT
wsC.merge_cells("B2:J2")
hdr_row = 4
cols = ["Year", "Order volume", "Auto-processed", "Manual", "Headcount\n(floor: no layoffs)",
        "Labor cost", "Subscription", "Implementation", "Maintenance", "Total cost"]
for c, h in enumerate(cols, start=2):
    hdr(wsC, hdr_row, c, h)

adoption_keys = ["Automation adoption — Year 1 (% of volume auto-processed)",
                  "Automation adoption — Year 2", "Automation adoption — Year 3"]
r = hdr_row + 1
for i, yr in enumerate([1, 2, 3]):
    wsC.cell(row=r, column=2, value=yr)
    wsC.cell(row=r, column=3, value=f"='Option B - Hire'!C{B_ROWS[0]+i}").font = GREEN
    wsC.cell(row=r, column=3).number_format = "0"
    wsC.cell(row=r, column=4, value=f"=C{r}*{A(adoption_keys[i])}").font = BLACK
    wsC.cell(row=r, column=4).number_format = "0"
    wsC.cell(row=r, column=5, value=f"=C{r}-D{r}").font = BLACK
    wsC.cell(row=r, column=5).number_format = "0"
    wsC.cell(row=r, column=6,
             value=f"=MAX({A('Current headcount (data-entry clerks)')},CEILING(E{r}/{A('Orders processed / clerk / year (manual)')},1))").font = BLACK
    wsC.cell(row=r, column=7, value=f"=F{r}*{A('Fully loaded cost / clerk / year')}").font = BLACK
    wsC.cell(row=r, column=7).number_format = "$#,##0"
    wsC.cell(row=r, column=8, value=f"={A('RPA/OCR platform subscription / year')}").font = BLACK
    wsC.cell(row=r, column=8).number_format = "$#,##0"
    wsC.cell(row=r, column=9, value=(f"={A('One-time implementation cost (Year 1 only)')}" if yr == 1 else 0))
    wsC.cell(row=r, column=9).font = BLACK
    wsC.cell(row=r, column=9).number_format = "$#,##0"
    wsC.cell(row=r, column=10, value=f"={A('Ongoing template/maintenance cost / year')}").font = BLACK
    wsC.cell(row=r, column=10).number_format = "$#,##0"
    wsC.cell(row=r, column=11, value=f"=G{r}+H{r}+I{r}+J{r}").font = BOLD
    wsC.cell(row=r, column=11).number_format = "$#,##0"
    for c in range(2, 12):
        wsC.cell(row=r, column=c).border = BORDER
    r += 1
C_ROWS = (hdr_row + 1, r - 1)
r += 1
wsC.cell(row=r, column=2, value="3-year total").font = BOLD
wsC.cell(row=r, column=11, value=f"=SUM(K{C_ROWS[0]}:K{C_ROWS[1]})").font = BOLD
wsC.cell(row=r, column=11).number_format = "$#,##0"
C_TOTAL_ROW = r
r += 2

wsC.cell(row=r, column=2, value="Error cost — Option C (auto share + manual share, blended rates)").font = BOLD
r += 1
hdr_row2 = r
for c, h in enumerate(["Year", "Auto-processed errors ($)", "Manual errors ($)", "Total error cost"], start=2):
    hdr(wsC, hdr_row2, c, h)
r += 1
for i, yr in enumerate([1, 2, 3]):
    src = C_ROWS[0] + i
    wsC.cell(row=r, column=2, value=yr)
    wsC.cell(row=r, column=3, value=f"=D{src}*{A('Automated (OCR-validated) order error rate')}*{A('Average cost per order error (rework, reship, CS time)')}").font = BLACK
    wsC.cell(row=r, column=3).number_format = "$#,##0"
    wsC.cell(row=r, column=4, value=f"=E{src}*{A('Manual order error rate (needs correction/rework)')}*{A('Average cost per order error (rework, reship, CS time)')}").font = BLACK
    wsC.cell(row=r, column=4).number_format = "$#,##0"
    wsC.cell(row=r, column=5, value=f"=C{r}+D{r}").font = BOLD
    wsC.cell(row=r, column=5).number_format = "$#,##0"
    for c in range(2, 6):
        wsC.cell(row=r, column=c).border = BORDER
    r += 1
ERR_C_ROWS = (hdr_row2 + 1, r - 1)

autosize(wsC, [3, 6, 13, 12, 12, 14, 12, 13, 15, 13, 13])

# ---------------------------------------------------------------- NPV & Payback
wsN = wb.create_sheet("NPV & Payback")
wsN["B2"] = "Net Benefit of Automating (Option C) vs. Hiring (Option B)"
wsN["B2"].font = TITLE_FONT
wsN.merge_cells("B2:H2")

hdr_row = 4
for c, h in enumerate(["Year", "Option B cost", "Option C cost", "Cost avoided (B-C)",
                        "Error cost — Option B (all-manual baseline)", "Error cost — Option C",
                        "Error cost avoided", "Net benefit"], start=2):
    hdr(wsN, hdr_row, c, h)
r = hdr_row + 1
for i, yr in enumerate([1, 2, 3]):
    b_row = B_ROWS[0] + i
    c_row = C_ROWS[0] + i
    err_row = ERR_C_ROWS[0] + i
    wsN.cell(row=r, column=2, value=yr)
    wsN.cell(row=r, column=3, value=f"='Option B - Hire'!H{b_row}").font = GREEN
    wsN.cell(row=r, column=3).number_format = "$#,##0"
    wsN.cell(row=r, column=4, value=f"='Option C - Automate'!K{c_row}").font = GREEN
    wsN.cell(row=r, column=4).number_format = "$#,##0"
    wsN.cell(row=r, column=5, value=f"=C{r}-D{r}").font = BLACK
    wsN.cell(row=r, column=5).number_format = "$#,##0"
    wsN.cell(row=r, column=6,
             value=f"='Option C - Automate'!C{c_row}*{A('Manual order error rate (needs correction/rework)')}*{A('Average cost per order error (rework, reship, CS time)')}").font = GREEN
    wsN.cell(row=r, column=6).number_format = "$#,##0"
    wsN.cell(row=r, column=7, value=f"='Option C - Automate'!E{err_row}").font = GREEN
    wsN.cell(row=r, column=7).number_format = "$#,##0"
    wsN.cell(row=r, column=8, value=f"=F{r}-G{r}").font = BLACK
    wsN.cell(row=r, column=8).number_format = "$#,##0"
    wsN.cell(row=r, column=9, value=f"=E{r}+H{r}").font = BOLD
    wsN.cell(row=r, column=9).number_format = "$#,##0"
    for c in range(2, 10):
        wsN.cell(row=r, column=c).border = BORDER
    r += 1
N_ROWS = (hdr_row + 1, r - 1)
r += 1
wsN.cell(row=r, column=2, value="3-year total net benefit").font = BOLD
wsN.cell(row=r, column=9, value=f"=SUM(I{N_ROWS[0]}:I{N_ROWS[1]})").font = BOLD
wsN.cell(row=r, column=9).number_format = "$#,##0"
N_TOTAL_ROW = r
r += 3

wsN.cell(row=r, column=2, value="Discounted cash flow (NPV)").font = BOLD
r += 1
hdr_row2 = r
for c, h in enumerate(["Year", "Net benefit", "Discount factor", "Present value", "Cumulative PV"], start=2):
    hdr(wsN, hdr_row2, c, h)
r += 1
PV_ROWS = []
for i, yr in enumerate([1, 2, 3]):
    src = N_ROWS[0] + i
    wsN.cell(row=r, column=2, value=yr)
    wsN.cell(row=r, column=3, value=f"=I{src}").font = GREEN
    wsN.cell(row=r, column=3).number_format = "$#,##0"
    wsN.cell(row=r, column=4, value=f"=1/(1+{A('Discount rate (WACC)')})^B{r}").font = BLACK
    wsN.cell(row=r, column=4).number_format = "0.000"
    wsN.cell(row=r, column=5, value=f"=C{r}*D{r}").font = BLACK
    wsN.cell(row=r, column=5).number_format = "$#,##0"
    if i == 0:
        wsN.cell(row=r, column=6, value=f"=E{r}").font = BLACK
    else:
        wsN.cell(row=r, column=6, value=f"=F{r-1}+E{r}").font = BLACK
    wsN.cell(row=r, column=6).number_format = "$#,##0"
    for c in range(2, 7):
        wsN.cell(row=r, column=c).border = BORDER
    PV_ROWS.append(r)
    r += 1
r += 1
wsN.cell(row=r, column=2, value="3-year NPV @ discount rate").font = BOLD
wsN.cell(row=r, column=5, value=f"=SUM(E{PV_ROWS[0]}:E{PV_ROWS[-1]})").font = BOLD
wsN.cell(row=r, column=5).number_format = "$#,##0"
NPV_ROW = r
r += 2

wsN.cell(row=r, column=2, value="Simple payback (undiscounted, using cumulative net benefit)").font = BOLD
r += 1
CUM_ROWS = []
for i, yr in enumerate([1, 2, 3]):
    src = N_ROWS[0] + i
    wsN.cell(row=r, column=2, value=f"Cumulative net benefit through Year {yr}")
    if i == 0:
        wsN.cell(row=r, column=5, value=f"=I{src}").font = BLACK
    else:
        wsN.cell(row=r, column=5, value=f"=E{r-1}+I{src}").font = BLACK
    wsN.cell(row=r, column=5).number_format = "$#,##0"
    CUM_ROWS.append(r)
    r += 1
r += 1
wsN.cell(row=r, column=2, value="Payback period (years)").font = BOLD
# find the year cumulative turns positive, interpolate within that year
payback_formula = (
    f'=IF(E{CUM_ROWS[0]}>=0,-E{CUM_ROWS[0]}/I{N_ROWS[0]}+1,'
    f'IF(E{CUM_ROWS[1]}>=0,1+(-E{CUM_ROWS[0]})/I{N_ROWS[0]+1},'
    f'IF(E{CUM_ROWS[2]}>=0,2+(-E{CUM_ROWS[1]})/I{N_ROWS[0]+2},"beyond 3 years")))'
)
wsN.cell(row=r, column=5, value=payback_formula).font = BOLD
wsN.cell(row=r, column=5).number_format = "0.00"
PAYBACK_ROW = r

autosize(wsN, [3, 40, 15, 15, 15, 15, 15, 15, 15])

# ---------------------------------------------------------------- Summary
wsS = wb.create_sheet("Summary")
wsS["B2"] = "Summary"
wsS["B2"].font = TITLE_FONT
wsS.merge_cells("B2:F2")
rows = [
    ("Option B — 3-year total cost (hire to keep pace)", f"='Option B - Hire'!H{B_TOTAL_ROW}", "$#,##0"),
    ("Option C — 3-year total cost (automate)", f"='Option C - Automate'!K{C_TOTAL_ROW}", "$#,##0"),
    ("3-year net benefit of automating (cost + error-avoidance)", f"='NPV & Payback'!I{N_TOTAL_ROW}", "$#,##0"),
    ("3-year NPV @ discount rate", f"='NPV & Payback'!E{NPV_ROW}", "$#,##0"),
    ("Payback period (years)", f"='NPV & Payback'!E{PAYBACK_ROW}", "0.00"),
]
rr = 4
for label, formula, fmt in rows:
    wsS.cell(row=rr, column=2, value=label)
    c = wsS.cell(row=rr, column=5, value=formula)
    c.font = GREEN
    c.number_format = fmt
    c.font = Font(name="Arial", size=11, bold=True, color="008000")
    for cc in range(2, 6):
        wsS.cell(row=rr, column=cc).border = BORDER
    rr += 1
autosize(wsS, [3, 50, 3, 3, 16])

wb.save(OUT)
print("Saved", OUT)
