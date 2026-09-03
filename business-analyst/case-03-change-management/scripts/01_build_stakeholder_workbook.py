"""
Builds stakeholder_analysis.xlsx for BA Case 03 (Change Management).

Sheets:
  Overview            - legend, scope
  Power-Interest Grid  - 12 stakeholders, formula-derived quadrant + engagement strategy
  RACI Matrix          - 9 integration activities x 7 roles, formula-validated (exactly one A/row)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "artifacts" / "stakeholder_analysis.xlsx"
OUT.parent.mkdir(exist_ok=True)

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
BOLD = Font(name="Arial", size=10, bold=True)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="44603B")
YELLOW = PatternFill("solid", fgColor="FFFF00")
QUAD_FILLS = {
    "Manage Closely": PatternFill("solid", fgColor="E4EBE1"),
    "Keep Satisfied": PatternFill("solid", fgColor="FCEFD9"),
    "Keep Informed": PatternFill("solid", fgColor="DCEAF7"),
    "Monitor": PatternFill("solid", fgColor="F1F1F1"),
}
THIN = Side(style="thin", color="B9C2C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="44603B")

wb = Workbook()

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c

# ---------------------------------------------------------------- Overview
ws = wb.active
ws.title = "Overview"
ws["B2"] = "Case 03 — Stakeholder Analysis"
ws["B2"].font = TITLE_FONT
ws["B4"] = ("Supports the Change Management Plan for integrating Colton Regional Supply's "
            "warehouse operations into Brightpath Distribution after acquisition. Two tabs: "
            "a power/interest grid (who to engage, and how) and a RACI matrix for the nine "
            "key integration activities.")
ws["B4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 48
ws["B6"] = "Color legend"
ws["B6"].font = BOLD
for i, (label, desc, font) in enumerate([
    ("Blue text", "Hardcoded input — power/interest rating or RACI assignment", BLUE),
    ("Yellow fill", "Key assumption", None),
    ("Black text", "Formula", BLACK),
], start=7):
    ws.cell(row=i, column=2, value=label).font = font or BOLD
    if font is None:
        ws.cell(row=i, column=2).fill = YELLOW
    ws.cell(row=i, column=3, value=desc)
autosize(ws, [3, 16, 70])

# ---------------------------------------------------------------- Power-Interest Grid
ws2 = wb.create_sheet("Power-Interest Grid")
ws2["B2"] = "Power / Interest Grid"
ws2["B2"].font = TITLE_FONT
ws2.merge_cells("B2:G2")
ws2["B4"] = "High threshold — a rating above this counts as \"high\" (1-5 scale)"
ws2["E4"] = 3
ws2["E4"].font = BLUE
ws2["E4"].fill = YELLOW

hdr_row = 6
for c, h in enumerate(["Stakeholder", "Power (1-5)", "Interest (1-5)", "Quadrant", "Engagement strategy"], start=2):
    hdr(ws2, hdr_row, c, h)

stakeholders = [
    ("CEO (Brightpath)", 5, 2, "Brief at milestones only; do not overload with operational detail — needs confidence the integration is on track, not the mechanics."),
    ("COO / Integration Sponsor", 5, 5, "Weekly steering check-in; owns final call on any scope or timeline change."),
    ("VP Operations (Brightpath)", 4, 5, "Co-owns the plan; weekly working session with the integration team."),
    ("Colton Regional GM", 4, 5, "Co-owns the plan; primary voice for Colton staff concerns — must be seen as a partner, not a subordinate, or the whole rollout reads as a takeover."),
    ("Brightpath Warehouse Ops Managers", 3, 4, "Involved in shift/scheduling and safety-procedure design; their buy-in determines floor-level adoption."),
    ("Colton Warehouse Ops Managers", 3, 5, "Same involvement as Brightpath counterparts, deliberately mirrored — asymmetric treatment here is a top resistance risk."),
    ("IT / Systems Integration Lead", 3, 3, "Standing biweekly sync; escalation path for data-migration issues."),
    ("HR Director", 4, 4, "Co-owns communications and the training plan; accountable for consistent messaging."),
    ("Finance / Payroll", 3, 3, "Consulted on payroll-system cutover timing only; not involved in operational design."),
    ("Brightpath Warehouse Staff", 2, 4, "Town halls + team-lead cascade; two-way channel for questions, not one-way announcements."),
    ("Colton Warehouse Staff", 2, 5, "Highest anxiety group (job security, new management) — most frequent, most concrete communication; specific answers, not general reassurance."),
    ("Key Colton-region customers", 3, 2, "Single proactive letter/call ahead of cutover confirming no service disruption; monitor for complaints during go-live week."),
]

r = hdr_row + 1
for name, power, interest, strategy in stakeholders:
    ws2.cell(row=r, column=2, value=name)
    ws2.cell(row=r, column=3, value=power).font = BLUE
    ws2.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    ws2.cell(row=r, column=4, value=interest).font = BLUE
    ws2.cell(row=r, column=4).alignment = Alignment(horizontal="center")
    formula = (f'=IF(AND(C{r}>$E$4,D{r}>$E$4),"Manage Closely",'
               f'IF(AND(C{r}>$E$4,D{r}<=$E$4),"Keep Satisfied",'
               f'IF(AND(C{r}<=$E$4,D{r}>$E$4),"Keep Informed","Monitor")))')
    ws2.cell(row=r, column=5, value=formula).font = BLACK
    ws2.cell(row=r, column=6, value=strategy).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(2, 7):
        ws2.cell(row=r, column=c).border = BORDER
    ws2.row_dimensions[r].height = 32
    r += 1
LASTROW_SH = r - 1

r += 1
ws2.cell(row=r, column=2, value="Count by quadrant").font = BOLD
r += 1
for q in ["Manage Closely", "Keep Satisfied", "Keep Informed", "Monitor"]:
    ws2.cell(row=r, column=2, value=q)
    ws2.cell(row=r, column=3, value=f'=COUNTIF(E{hdr_row+1}:E{LASTROW_SH},B{r})').font = BLACK
    r += 1

autosize(ws2, [3, 32, 12, 12, 16, 62])

# ---------------------------------------------------------------- RACI Matrix
ws3 = wb.create_sheet("RACI Matrix")
ws3["B2"] = "RACI Matrix — Integration Activities"
ws3["B2"].font = TITLE_FONT
ws3.merge_cells("B2:J2")
ws3["B4"] = "R = Responsible · A = Accountable (exactly one per activity) · C = Consulted · I = Informed"
ws3.merge_cells("B4:J4")

roles = ["COO / Sponsor", "VP Ops (Brightpath)", "Colton GM", "Warehouse Ops Managers",
         "IT Integration Lead", "HR Director", "Warehouse Staff"]
hdr_row = 6
hdr(ws3, hdr_row, 2, "Activity")
for i, role in enumerate(roles):
    hdr(ws3, hdr_row, 3 + i, role)
hdr(ws3, hdr_row, 3 + len(roles), "Check")

activities = [
    ("Approve unified WMS platform selection", ["A", "C", "C", "C", "R", "I", ""]),
    ("Design unified shift/scheduling policy", ["I", "A", "C", "R", "", "C", "I"]),
    ("Communicate role/reporting-line changes to staff", ["I", "C", "C", "R", "", "A", "I"]),
    ("Migrate Colton inventory data to unified WMS", ["I", "I", "C", "R", "A", "", ""]),
    ("Deliver cross-training on new WMS & procedures", ["", "I", "I", "A", "R", "C", "I"]),
    ("Set unified safety & incident-reporting procedures", ["I", "A", "C", "R", "", "C", "I"]),
    ("Finalize integrated org chart & reporting lines", ["A", "C", "C", "I", "", "R", "I"]),
    ("Go-live cutover weekend execution", ["I", "A", "C", "R", "C", "", "I"]),
    ("Post-integration 30-day review", ["I", "A", "C", "C", "C", "C", ""]),
]

r = hdr_row + 1
for name, raci in activities:
    ws3.cell(row=r, column=2, value=name)
    for i, val in enumerate(raci):
        cell = ws3.cell(row=r, column=3 + i, value=val if val else None)
        cell.font = BLUE
        cell.alignment = Alignment(horizontal="center")
    check_col = 3 + len(roles)
    rng_start = get_column_letter(3)
    rng_end = get_column_letter(3 + len(roles) - 1)
    formula = f'=IF(COUNTIF({rng_start}{r}:{rng_end}{r},"A")=1,"OK","CHECK")'
    ws3.cell(row=r, column=check_col, value=formula).font = BLACK
    for c in range(2, check_col + 1):
        ws3.cell(row=r, column=c).border = BORDER
    r += 1
LASTROW_RACI = r - 1

r += 1
ws3.cell(row=r, column=2, value="All activities have exactly one Accountable?").font = BOLD
ws3.cell(row=r, column=3, value=f'=IF(COUNTIF(J{hdr_row+1}:J{LASTROW_RACI},"CHECK")=0,"Yes — validated","No — fix flagged rows")').font = BLACK

autosize(ws3, [3, 40, 13, 15, 12, 16, 14, 12, 13, 10])

wb.save(OUT)
print("Saved", OUT)
