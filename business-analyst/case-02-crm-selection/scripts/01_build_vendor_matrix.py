"""
Builds vendor_evaluation_matrix.xlsx for BA Case 02 (CRM selection).

Sheets:
  Overview     - legend, scope
  Scoring      - weighted criteria x 3 vendors, all formula-driven
  TCO Detail   - 3-year total cost of ownership build-up per vendor, feeds
                 the Scoring sheet's TCO score via a formula (not typed in)
"""
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "artifacts" / "vendor_evaluation_matrix.xlsx"
OUT.parent.mkdir(exist_ok=True)

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10, color="000000")
GREEN = Font(name="Arial", size=10, color="008000")
BOLD = Font(name="Arial", size=10, bold=True)
HDR_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="44603B")
WIN_FILL = PatternFill("solid", fgColor="E4EBE1")
YELLOW = PatternFill("solid", fgColor="FFFF00")
THIN = Side(style="thin", color="B9C2C0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="44603B")

wb = Workbook()

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def hdr(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER
    return c

VENDORS = ["Pinnacle Sales Cloud", "Northstar CRM", "Vertex Sales Suite"]

# ---------------------------------------------------------------- Overview
ws = wb.active
ws.title = "Overview"
ws["B2"] = "Case 02 — CRM Vendor Evaluation Matrix"
ws["B2"].font = TITLE_FONT
ws["B4"] = ("Supports the Business Requirements Document. Three vendors scored against "
            "weighted criteria drawn directly from the requirements-gathering interviews. "
            "Vendor names are fictional — this is a simulated evaluation for a portfolio "
            "case study, not a real product comparison.")
ws["B4"].alignment = Alignment(wrap_text=True)
ws.merge_cells("B4:H4")
ws.row_dimensions[4].height = 48
ws["B6"] = "Color legend"
ws["B6"].font = BOLD
legend = [("Blue text", "Hardcoded input — interview-informed score or a quoted vendor price", BLUE),
          ("Yellow fill", "Key assumption", None),
          ("Black text", "Formula", BLACK),
          ("Green text", "Link from another sheet", GREEN)]
r = 7
for label, desc, font in legend:
    ws.cell(row=r, column=2, value=label).font = font or BOLD
    if font is None:
        ws.cell(row=r, column=2).fill = YELLOW
    ws.cell(row=r, column=3, value=desc)
    r += 1
autosize(ws, [3, 16, 70])

# ---------------------------------------------------------------- TCO Detail
ws_tco = wb.create_sheet("TCO Detail")
ws_tco["B2"] = "3-Year Total Cost of Ownership"
ws_tco["B2"].font = TITLE_FONT
ws_tco.merge_cells("B2:F2")

ws_tco["B4"] = "Seat count"
ws_tco["E4"] = 40
ws_tco["E4"].font = BLUE
ws_tco["E4"].fill = YELLOW
ws_tco["B5"] = "Contract term (months)"
ws_tco["E5"] = 36
ws_tco["E5"].font = BLUE
ws_tco["E5"].fill = YELLOW

hdr_row = 7
for c, h in enumerate(["Vendor", "$/seat/month", "Implementation (one-time)",
                        "Data migration (one-time)", "License cost (3yr)", "3-Year TCO"], start=2):
    hdr(ws_tco, hdr_row, c, h)

vendor_pricing = [
    ("Pinnacle Sales Cloud", 145, 65000, 15000),
    ("Northstar CRM", 99, 42000, 22000),
    ("Vertex Sales Suite", 120, 38000, 9000),
]
r = hdr_row + 1
TCO_ROWS = {}
for name, seat_price, impl, migr in vendor_pricing:
    ws_tco.cell(row=r, column=2, value=name).font = BLUE
    ws_tco.cell(row=r, column=3, value=seat_price).font = BLUE
    ws_tco.cell(row=r, column=3).number_format = "$#,##0"
    ws_tco.cell(row=r, column=4, value=impl).font = BLUE
    ws_tco.cell(row=r, column=4).number_format = "$#,##0"
    ws_tco.cell(row=r, column=5, value=migr).font = BLUE
    ws_tco.cell(row=r, column=5).number_format = "$#,##0"
    ws_tco.cell(row=r, column=6, value=f"=C{r}*$E$4*$E$5").font = BLACK
    ws_tco.cell(row=r, column=6).number_format = "$#,##0"
    ws_tco.cell(row=r, column=7, value=f"=F{r}+D{r}+E{r}").font = BOLD
    ws_tco.cell(row=r, column=7).number_format = "$#,##0"
    for c in range(2, 8):
        ws_tco.cell(row=r, column=c).border = BORDER
    TCO_ROWS[name] = r
    r += 1
TCO_MIN_ROW = r  # blank row then min/max
r += 1
ws_tco.cell(row=r, column=2, value="Lowest 3-yr TCO").font = BOLD
ws_tco.cell(row=r, column=7, value=f"=MIN(G{hdr_row+1}:G{hdr_row+3})").font = BLACK
ws_tco.cell(row=r, column=7).number_format = "$#,##0"
TCO_MIN_CELL = f"G{r}"
r += 1
ws_tco.cell(row=r, column=2, value="Highest 3-yr TCO").font = BOLD
ws_tco.cell(row=r, column=7, value=f"=MAX(G{hdr_row+1}:G{hdr_row+3})").font = BLACK
ws_tco.cell(row=r, column=7).number_format = "$#,##0"
TCO_MAX_CELL = f"G{r}"

autosize(ws_tco, [3, 24, 15, 22, 20, 16, 16])

# ---------------------------------------------------------------- Scoring
ws_s = wb.create_sheet("Scoring")
ws_s["B2"] = "Weighted Vendor Scoring"
ws_s["B2"].font = TITLE_FONT
ws_s.merge_cells("B2:H2")
ws_s["B4"] = ("Each criterion scored 1 (poor) to 5 (excellent) by the evaluation panel "
              "(Sales Ops, IT, two Sales Managers) during vendor demos, except Total Cost "
              "of Ownership, which is derived automatically from the TCO Detail sheet — "
              "cheapest option scores 5, most expensive scores 1, scaled linearly between.")
ws_s.merge_cells("B4:H4")
ws_s["B4"].alignment = Alignment(wrap_text=True)
ws_s.row_dimensions[4].height = 34

hdr_row = 6
cols = ["Criterion", "Weight"] + VENDORS
for c, h in enumerate(cols, start=2):
    hdr(ws_s, hdr_row, c, h)

criteria = [
    ("Pipeline & opportunity management", 0.20, [4, 5, 3]),
    ("Reporting & forecasting", 0.18, [5, 4, 3]),
    ("Mobile usability", 0.15, [3, 5, 4]),
    ("Integration (ERP, email/calendar, marketing)", 0.15, [4, 3, 5]),
    ("Data migration & onboarding effort (ease)", 0.10, [3, 4, 5]),
    ("Vendor support & SLA", 0.10, [4, 3, 5]),
]
r = hdr_row + 1
CRIT_ROWS = []
for label, weight, scores in criteria:
    ws_s.cell(row=r, column=2, value=label)
    ws_s.cell(row=r, column=3, value=weight).font = BLUE
    ws_s.cell(row=r, column=3).number_format = "0%"
    ws_s.cell(row=r, column=3).fill = YELLOW
    for i, sc in enumerate(scores):
        cell = ws_s.cell(row=r, column=4 + i, value=sc)
        cell.font = BLUE
        cell.alignment = Alignment(horizontal="center")
    for c in range(2, 7):
        ws_s.cell(row=r, column=c).border = BORDER
    CRIT_ROWS.append(r)
    r += 1

TCO_ROW = r
ws_s.cell(row=r, column=2, value="Total cost of ownership (3-yr, lower is better)")
ws_s.cell(row=r, column=3, value=0.12).font = BLUE
ws_s.cell(row=r, column=3).number_format = "0%"
ws_s.cell(row=r, column=3).fill = YELLOW
for i, name in enumerate(VENDORS):
    trow = TCO_ROWS[name]
    formula = f"=5-((\'TCO Detail\'!G{trow}-\'TCO Detail\'!{TCO_MIN_CELL})/(\'TCO Detail\'!{TCO_MAX_CELL}-\'TCO Detail\'!{TCO_MIN_CELL}))*4"
    cell = ws_s.cell(row=r, column=4 + i, value=formula)
    cell.font = GREEN
    cell.number_format = "0.00"
for c in range(2, 7):
    ws_s.cell(row=r, column=c).border = BORDER
CRIT_ROWS.append(r)
r += 1

WEIGHT_CHECK_ROW = r
ws_s.cell(row=r, column=2, value="Weight check (should be 100%)").font = BOLD
ws_s.cell(row=r, column=3, value=f"=SUM(C{hdr_row+1}:C{TCO_ROW})").font = BLACK
ws_s.cell(row=r, column=3).number_format = "0%"
r += 2

WTOTAL_ROW = r
ws_s.cell(row=r, column=2, value="Weighted total").font = BOLD
for i in range(3):
    col = get_column_letter(4 + i)
    parts = "+".join(f"{col}{cr}*$C{cr}" for cr in CRIT_ROWS)
    cell = ws_s.cell(row=r, column=4 + i, value=f"={parts}")
    cell.font = BOLD
    cell.number_format = "0.00"
    cell.border = BORDER
ws_s.cell(row=r, column=2).border = BORDER
ws_s.cell(row=r, column=3).border = BORDER
r += 1

RANK_ROW = r
ws_s.cell(row=r, column=2, value="Rank").font = BOLD
for i in range(3):
    col = get_column_letter(4 + i)
    wcol_range = f"D{WTOTAL_ROW}:F{WTOTAL_ROW}"
    cell = ws_s.cell(row=r, column=4 + i, value=f"=RANK({col}{WTOTAL_ROW},{wcol_range})")
    cell.font = BLACK
    cell.alignment = Alignment(horizontal="center")
r += 2

ws_s.cell(row=r, column=2, value="Recommended vendor").font = BOLD
ws_s.cell(row=r, column=4,
          value=f'=INDEX(D{hdr_row}:F{hdr_row},MATCH(1,D{RANK_ROW}:F{RANK_ROW},0))').font = BOLD
ws_s.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
ws_s.cell(row=r, column=4).fill = WIN_FILL

autosize(ws_s, [3, 44, 10, 22, 22, 22])

wb.save(OUT)
print("Saved", OUT)
