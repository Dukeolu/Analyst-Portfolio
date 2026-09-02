"""
Case 04 -- build the Excel workbook deliverable.

Headcount and attrition-rate reporting is classic Excel territory in a
real HR/People Ops team, so -- as with Case 03's Finance workbook -- this
is a genuine live-formula model, not a values dump.

Sheets:
  README                 context + color legend
  Headcount & Attrition  144-row monthly roll-forward by department (Beginning HC ->
                         Hires -> Terms -> Ending HC -> annualized attrition rate)
  Recruiting Funnel      channel-level conversion rates (SUMIFS) + the Warehouse & Ops
                         x Job Board bottleneck spotlighted, with conditional formatting
  Cost of the Leak       assumption-driven $ impact of early attrition and funnel waste,
                         all formula-derived from the two tables above
"""
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[1]
RAW = BASE / "data" / "raw"
PROCESSED = BASE / "data" / "processed"
OUT = BASE / "excel" / "case04_hr_analysis.xlsx"

hc = pd.read_csv(PROCESSED / "monthly_headcount_by_dept.csv")
funnel = pd.read_csv(RAW / "recruiting_funnel.csv")
emp = pd.read_csv(RAW / "employees.csv")

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1F2933")
HEAD_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
INPUT_FONT = Font(name=FONT, color="0000FF", size=10)
FORMULA_FONT = Font(name=FONT, color="000000", size=10)
ASSUMPTION_FILL = PatternFill("solid", fgColor="FFFF00")
TITLE_FONT = Font(name=FONT, bold=True, size=15)
SUB_FONT = Font(name=FONT, size=10.5, italic=True, color="52514E")
MONEY = '$#,##0;($#,##0);"-"'
PCT1 = '0.0%'

wb = Workbook()

# =========================================================== README
ws = wb.active
ws.title = "README"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 100
ws["A1"] = "Case 04 — Where the Hiring Funnel (and Headcount) Leaks"
ws["A1"].font = TITLE_FONT
ws["A2"] = "HR & People Operations case study — recruiting funnel efficiency and voluntary attrition"
ws["A2"].font = SUB_FONT
lines = [
    "",
    "Business problem: attrition is above plan and recruiting can't say whether the bottleneck is",
    "who's being hired, where they sit, or what happens to them in year one. This workbook traces",
    "both the hiring funnel and post-hire attrition back to the same root cause.",
    "",
    "Sheet guide:",
    "  Headcount & Attrition   144-row monthly roll-forward by department (Beginning HC -> Hires ->",
    "                          Terminations -> Ending HC), with an annualized attrition rate formula.",
    "                          Starting headcounts (yellow) are the one true assumption in this sheet.",
    "  Recruiting Funnel       Applied -> Screened -> Interviewed -> Offered -> Hired conversion by",
    "                          channel (SUMIFS), plus the Warehouse & Ops x Job Board bottleneck.",
    "  Cost of the Leak        Formula-driven $ impact of early (<=3 month) attrition by channel and",
    "                          of the screening bottleneck, with every assumption in a labeled, editable cell.",
    "",
    "Color legend:",
    "  Blue text    = input data (as pulled from HRIS/ATS systems)",
    "  Yellow fill  = key assumptions -- edit these and the workbook recalculates",
    "  Black text   = formulas",
    "",
    "Data note: this dataset is simulated (see the case README for why and how) -- no demographic",
    "attributes (age, gender, etc.) are modeled. Attrition risk here is driven entirely by structural",
    "factors: department, comp-to-market ratio, overtime load, manager span, engagement, and hiring",
    "channel. Two deliberate mechanisms are seeded: a screening backlog specific to Warehouse & Ops'",
    "Job Board pipeline, and elevated early-tenure attrition for Job Board hires network-wide.",
]
for i, line in enumerate(lines, start=3):
    ws[f"A{i}"] = line
    ws[f"A{i}"].font = Font(name=FONT, size=10.5, bold=line.strip().endswith(":") and not line.startswith(" "))

# =========================================================== Headcount & Attrition
ws2 = wb.create_sheet("Headcount & Attrition")
headers = ["Month", "Department", "Beginning HC", "Hires", "Terminations", "Ending HC", "Avg HC", "Annualized Attrition Rate"]
for c, h in enumerate(headers, start=1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
ws2.freeze_panes = "A2"

departments = sorted(hc["department"].unique())
START_HC = {  # assumption: headcount at the start of the observation window (Dec 2023)
    "Sales": 175, "Customer Support": 145, "Warehouse & Ops": 230,
    "Engineering": 135, "Finance & Accounting": 105, "Marketing": 90,
}

# assumptions block
ws2.cell(row=1, column=10, value="Starting HC (Dec 2023)").font = HEAD_FONT
ws2.cell(row=1, column=10).fill = HEAD_FILL
for i, dept in enumerate(departments, start=2):
    ws2.cell(row=i, column=9, value=dept).font = FORMULA_FONT
    cell = ws2.cell(row=i, column=10, value=START_HC[dept])
    cell.font = INPUT_FONT
    cell.fill = ASSUMPTION_FILL
ws2.column_dimensions["I"].width = 20
ws2.column_dimensions["J"].width = 20

row = 2
dept_first_row = {}
for dept in departments:
    sub = hc[hc.department == dept].sort_values("month")
    for k, r in enumerate(sub.itertuples(index=False)):
        ws2.cell(row=row, column=1, value=r.month).font = INPUT_FONT
        ws2.cell(row=row, column=2, value=r.department).font = INPUT_FONT
        if k == 0:
            dept_row_idx = departments.index(dept) + 2
            ws2.cell(row=row, column=3, value=f"=J{dept_row_idx}").font = FORMULA_FONT
            dept_first_row[dept] = row
        else:
            ws2.cell(row=row, column=3, value=f"=F{row-1}").font = FORMULA_FONT
        ws2.cell(row=row, column=4, value=r.hires).font = INPUT_FONT
        ws2.cell(row=row, column=5, value=r.terminations).font = INPUT_FONT
        ws2.cell(row=row, column=6, value=f"=C{row}+D{row}-E{row}").font = FORMULA_FONT
        ws2.cell(row=row, column=7, value=f"=AVERAGE(C{row},F{row})").font = FORMULA_FONT
        ws2.cell(row=row, column=8, value=f"=IFERROR(E{row}/G{row}*12,0)").font = FORMULA_FONT
        ws2.cell(row=row, column=8).number_format = PCT1
        for col in (3, 6, 7):
            ws2.cell(row=row, column=col).number_format = '#,##0'
        row += 1
last_hc_row = row - 1
widths2 = [9, 20, 13, 8, 13, 11, 10, 20]
for c, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(c)].width = w

# =========================================================== Recruiting Funnel
ws3 = wb.create_sheet("Recruiting Funnel")
ws3.sheet_view.showGridLines = False
ws3["A1"] = "Recruiting Funnel Conversion by Channel (24 months, all departments)"
ws3["A1"].font = Font(name=FONT, bold=True, size=13)
ws3["A2"] = "Flat input table (below) drives the conversion-rate summary via SUMIFS."
ws3["A2"].font = SUB_FONT

fh = ["Month", "Department", "Channel", "Applied", "Screened", "Interviewed", "Offered", "Hired"]
for c, h in enumerate(fh, start=1):
    cell = ws3.cell(row=4, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
frow = 5
for r in funnel.itertuples(index=False):
    ws3.cell(row=frow, column=1, value=r.month).font = INPUT_FONT
    ws3.cell(row=frow, column=2, value=r.department).font = INPUT_FONT
    ws3.cell(row=frow, column=3, value=r.channel).font = INPUT_FONT
    ws3.cell(row=frow, column=4, value=r.applied).font = INPUT_FONT
    ws3.cell(row=frow, column=5, value=r.screened).font = INPUT_FONT
    ws3.cell(row=frow, column=6, value=r.interviewed).font = INPUT_FONT
    ws3.cell(row=frow, column=7, value=r.offered).font = INPUT_FONT
    ws3.cell(row=frow, column=8, value=r.hired).font = INPUT_FONT
    frow += 1
funnel_last_row = frow - 1
for c, w in zip(range(1, 9), [9, 20, 18, 9, 10, 11, 9, 8]):
    ws3.column_dimensions[get_column_letter(c)].width = w

summary_top = funnel_last_row + 3
ws3.cell(row=summary_top, column=1, value="Conversion Rate by Channel").font = Font(name=FONT, bold=True, size=12)
sh = ["Channel", "Applied", "Screened", "Interviewed", "Offered", "Hired", "Screen Rate", "Interview Rate", "Offer Rate", "Accept Rate"]
hrow = summary_top + 1
for c, h in enumerate(sh, start=1):
    cell = ws3.cell(row=hrow, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL

channels = sorted(funnel["channel"].unique())
for i, ch in enumerate(channels, start=hrow + 1):
    ws3.cell(row=i, column=1, value=ch).font = FORMULA_FONT
    for j, col in enumerate(["D", "E", "F", "G", "H"], start=2):
        ws3.cell(row=i, column=j,
                 value=f"=SUMIFS('Recruiting Funnel'!${col}$5:${col}${funnel_last_row},'Recruiting Funnel'!$C$5:$C${funnel_last_row},A{i})").font = FORMULA_FONT
    ws3.cell(row=i, column=7, value=f"=IFERROR(C{i}/B{i},0)").font = FORMULA_FONT
    ws3.cell(row=i, column=8, value=f"=IFERROR(D{i}/C{i},0)").font = FORMULA_FONT
    ws3.cell(row=i, column=9, value=f"=IFERROR(E{i}/D{i},0)").font = FORMULA_FONT
    ws3.cell(row=i, column=10, value=f"=IFERROR(F{i}/E{i},0)").font = FORMULA_FONT
    for col in (7, 8, 9, 10):
        ws3.cell(row=i, column=col).number_format = PCT1
channel_summary_last = hrow + len(channels)
ws3.conditional_formatting.add(
    f"G{hrow+1}:J{channel_summary_last}",
    ColorScaleRule(start_type="min", start_color="C0392B", mid_type="percentile", mid_value=50, mid_color="FFFFFF", end_type="max", end_color="1E8449"),
)

# spotlight table: Warehouse & Ops x Job Board vs Job Board overall
spot_top = channel_summary_last + 3
ws3.cell(row=spot_top, column=1, value="Spotlight: Warehouse & Ops x Job Board Bottleneck").font = Font(name=FONT, bold=True, size=12)
spot_hrow = spot_top + 1
for c, h in enumerate(["Segment", "Screened", "Interviewed", "Interview Rate"], start=1):
    cell = ws3.cell(row=spot_hrow, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
segs = [
    ("Warehouse & Ops x Job Board", "AND($B$5:$B$"+str(funnel_last_row)+"=\"Warehouse & Ops\",$C$5:$C$"+str(funnel_last_row)+"=\"Job Board\")"),
]
r = spot_hrow + 1
ws3.cell(row=r, column=1, value="Warehouse & Ops x Job Board").font = FORMULA_FONT
ws3.cell(row=r, column=2, value=f"=SUMIFS($E$5:$E${funnel_last_row},$B$5:$B${funnel_last_row},\"Warehouse & Ops\",$C$5:$C${funnel_last_row},\"Job Board\")").font = FORMULA_FONT
ws3.cell(row=r, column=3, value=f"=SUMIFS($F$5:$F${funnel_last_row},$B$5:$B${funnel_last_row},\"Warehouse & Ops\",$C$5:$C${funnel_last_row},\"Job Board\")").font = FORMULA_FONT
ws3.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)").font = FORMULA_FONT
r += 1
ws3.cell(row=r, column=1, value="Job Board, all departments").font = FORMULA_FONT
ws3.cell(row=r, column=2, value=f"=SUMIFS($E$5:$E${funnel_last_row},$C$5:$C${funnel_last_row},\"Job Board\")").font = FORMULA_FONT
ws3.cell(row=r, column=3, value=f"=SUMIFS($F$5:$F${funnel_last_row},$C$5:$C${funnel_last_row},\"Job Board\")").font = FORMULA_FONT
ws3.cell(row=r, column=4, value=f"=IFERROR(C{r}/B{r},0)").font = FORMULA_FONT
for rr in (spot_hrow + 1, spot_hrow + 2):
    ws3.cell(row=rr, column=4).number_format = PCT1

# =========================================================== Cost of the Leak
ws4 = wb.create_sheet("Cost of the Leak")
ws4.sheet_view.showGridLines = False
ws4["A1"] = "Cost of the Leak"
ws4["A1"].font = Font(name=FONT, bold=True, size=13)
ws4["A2"] = "Every input below is a labeled, editable assumption (yellow) -- change one and the totals recalculate."
ws4["A2"].font = SUB_FONT

# early attrition counts by channel (<=3 months tenure), computed from employees.csv
early = emp[emp["terminated_voluntary"] == "Yes"].copy()
early_counts = early[early["tenure_months"] <= 3].groupby("channel").size().reindex(channels, fill_value=0)
hires_by_channel = emp.groupby("channel").size().reindex(channels, fill_value=0)

ws4["A4"] = "Early (<=3mo) voluntary attrition by channel"
ws4["A4"].font = Font(name=FONT, bold=True, size=11)
eh = ["Channel", "Hires (24mo)", "Early Departures", "Early Attrition Rate", "Replacement Cost / Departure", "Cost of Early Attrition"]
for c, h in enumerate(eh, start=1):
    cell = ws4.cell(row=5, column=c, value=h)
    cell.font = HEAD_FONT
    cell.fill = HEAD_FILL
REPLACEMENT_COST = 9000
for i, ch in enumerate(channels, start=6):
    ws4.cell(row=i, column=1, value=ch).font = FORMULA_FONT
    ws4.cell(row=i, column=2, value=int(hires_by_channel[ch])).font = INPUT_FONT
    ws4.cell(row=i, column=3, value=int(early_counts[ch])).font = INPUT_FONT
    ws4.cell(row=i, column=4, value=f"=IFERROR(C{i}/B{i},0)").font = FORMULA_FONT
    ws4.cell(row=i, column=4).number_format = PCT1
    cell = ws4.cell(row=i, column=5, value=REPLACEMENT_COST)
    cell.font = INPUT_FONT
    cell.fill = ASSUMPTION_FILL
    cell.number_format = MONEY
    ws4.cell(row=i, column=6, value=f"=C{i}*E{i}").font = FORMULA_FONT
    ws4.cell(row=i, column=6).number_format = MONEY
early_last = 5 + len(channels)
ws4.cell(row=early_last + 1, column=1, value="Total").font = Font(name=FONT, bold=True)
ws4.cell(row=early_last + 1, column=3, value=f"=SUM(C6:C{early_last})").font = Font(name=FONT, bold=True)
ws4.cell(row=early_last + 1, column=6, value=f"=SUM(F6:F{early_last})").font = Font(name=FONT, bold=True)
ws4.cell(row=early_last + 1, column=6).number_format = MONEY

scenario_top = early_last + 4
ws4.cell(row=scenario_top, column=1, value="Scenario: Job Board's early-attrition rate matches Employee Referral's").font = Font(name=FONT, bold=True, size=11)
jb_row = 6 + channels.index("Job Board")
ref_row = 6 + channels.index("Employee Referral")
ws4.cell(row=scenario_top + 1, column=1, value="Job Board hires (24mo)").font = FORMULA_FONT
ws4.cell(row=scenario_top + 1, column=2, value=f"=B{jb_row}").font = FORMULA_FONT
ws4.cell(row=scenario_top + 2, column=1, value="Employee Referral early-attrition rate (target)").font = FORMULA_FONT
ws4.cell(row=scenario_top + 2, column=2, value=f"=D{ref_row}").font = FORMULA_FONT
ws4.cell(row=scenario_top + 2, column=2).number_format = PCT1
ws4.cell(row=scenario_top + 3, column=1, value="Implied early departures at target rate").font = FORMULA_FONT
ws4.cell(row=scenario_top + 3, column=2, value=f"=B{scenario_top+1}*B{scenario_top+2}").font = FORMULA_FONT
ws4.cell(row=scenario_top + 4, column=1, value="Actual Job Board early departures").font = FORMULA_FONT
ws4.cell(row=scenario_top + 4, column=2, value=f"=C{jb_row}").font = FORMULA_FONT
ws4.cell(row=scenario_top + 5, column=1, value="Departures avoided").font = Font(name=FONT, bold=True)
ws4.cell(row=scenario_top + 5, column=2, value=f"=B{scenario_top+4}-B{scenario_top+3}").font = Font(name=FONT, bold=True)
ws4.cell(row=scenario_top + 6, column=1, value="Annual savings (24mo total / 2)").font = Font(name=FONT, bold=True)
ws4.cell(row=scenario_top + 6, column=2, value=f"=B{scenario_top+5}*E{jb_row}/2").font = Font(name=FONT, bold=True)
ws4.cell(row=scenario_top + 6, column=2).number_format = MONEY
ws4.column_dimensions["A"].width = 46
for c, w in zip(range(2, 7), [14, 16, 18, 22, 20]):
    ws4.column_dimensions[get_column_letter(c)].width = w

chart = BarChart()
chart.type = "col"
chart.title = "Early Attrition Rate by Hiring Channel"
chart.y_axis.title = "Rate"
data = Reference(ws4, min_col=4, min_row=5, max_row=early_last)
cats = Reference(ws4, min_col=1, min_row=6, max_row=early_last)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.width, chart.height = 16, 9
ws4.add_chart(chart, "H5")

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"Workbook written to {OUT}")
